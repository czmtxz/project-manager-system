# -*- coding: utf-8 -*-
"""
项目费用归集系统 - 主程序
功能：项目管理、合同管理、发票管理、付款管理、往来款登记、图片OCR识别、Excel导入
"""

import os
import json
import sqlite3
import hashlib
import shutil
import threading
import schedule
import time
import re
import uuid
from datetime import datetime, timedelta
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, send_file, send_from_directory, flash, g)
from werkzeug.security import generate_password_hash, check_password_hash
from project_category_utils import (
    ensure_project_categories_table,
    fetch_leaf_categories,
    get_grouped_categories_json,
    get_project_categories_list,
    get_project_enabled_category_ids,
    project_uses_category_whitelist,
    set_project_categories,
    validate_project_category,
    category_ids_sql_filter,
)
from auth_utils import (
    ROLE_ADMIN,
    ROLE_CLIENT_COLLAB,
    ROLE_LABELS,
    MODULE_CLIENT_PORTAL,
    module_required,
    register_auth_hooks,
    login_redirect_for_role,
)
from client_portal_utils import (
    ensure_schema_extensions,
    resolve_or_create_customer,
    sales_order_customer_filter,
    assert_sales_order_access,
    company_scope_client_ids,
    aggregate_company_balance,
    build_portal_transactions,
    sync_deductions_for_customer,
    CLIENT_STATUS_APPROVED,
    CLIENT_STATUS_PENDING,
    CLIENT_STATUS_REJECTED,
    CLIENT_STATUS_DISABLED,
    normalize_client_status,
    get_client_by_id,
)
from client_collab_scope import (
    assert_customer_access,
    assign_customer_to_user,
    apply_client_account_scope_sql,
    apply_client_id_scope_sql,
    can_access_customer,
    ensure_collab_scope_schema,
    get_user_assignments,
    list_company_summaries_scoped,
    set_user_assignments,
)
from client_collab_ops import (
    PAYMENT_METHODS,
    list_company_summaries,
    get_company_workspace,
    record_client_recharge,
    record_client_outbound,
    parse_collab_excel,
    import_recharge_rows,
    import_outbound_rows,
    import_standard_excel_bundle,
    ocr_rows_for_recharge,
    ocr_rows_for_outbound,
    save_upload_file,
    primary_client_for_company,
)

STAFF_ROLES = [
    {'code': 'admin', 'name': '系统管理员'},
    {'code': 'client_collab', 'name': '客户协同专员'},
    {'code': 'finance', 'name': '财务'},
    {'code': 'manager', 'name': '项目经理'},
    {'code': 'user', 'name': '普通用户'},
]

app = Flask(__name__)
app.secret_key = 'project_manager_secret_key_2024'
app.config['DATABASE'] = 'project_manager.db'
app.config['BACKUP_DIR'] = 'backups'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 8 * 1024 * 1024  # 8MB

# 确保目录存在
os.makedirs(app.config['BACKUP_DIR'], exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'ocr'), exist_ok=True)


# ==================== 数据库工具 ====================

def get_db():
    """获取数据库连接"""
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        if not app.config.get('_schema_extended'):
            ensure_schema_extensions(g.db)
            ensure_collab_scope_schema(g.db)
            from report_hub_prefs import ensure_report_hub_prefs_schema
            ensure_report_hub_prefs_schema(g.db)
            app.config['_schema_extended'] = True
    return g.db


@app.teardown_appcontext
def close_db(exception):
    """关闭数据库连接"""
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    """初始化数据库"""
    db = sqlite3.connect(app.config['DATABASE'])
    db.row_factory = sqlite3.Row
    cursor = db.cursor()

    # 用户表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 项目表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            status TEXT DEFAULT '进行中',
            budget REAL DEFAULT 0,
            start_date DATE,
            end_date DATE,
            user_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    # 参与人表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS participants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            role TEXT DEFAULT 'member',
            status TEXT DEFAULT 'active',
            remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 项目参与人关联表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_participants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            participant_id INTEGER,
            project_role TEXT DEFAULT 'member',
            investment_ratio REAL DEFAULT 0,
            dividend_ratio REAL DEFAULT 0,
            joined_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (participant_id) REFERENCES participants(id)
        )
    ''')
    # 同一项目下同一参与人唯一，保证 INSERT OR REPLACE 正常工作、避免重复行
    cursor.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_pp_project_participant
        ON project_participants(project_id, participant_id)
    ''')

    # 参与人账户表（用于OCR匹配）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS participant_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            participant_id INTEGER,
            account_type TEXT,
            account_name TEXT,
            account_tail TEXT,
            keywords TEXT,
            is_default INTEGER DEFAULT 0,
            FOREIGN KEY (participant_id) REFERENCES participants(id)
        )
    ''')

    # 合同表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            contract_no TEXT,
            contract_name TEXT,
            contract_type TEXT,
            amount REAL DEFAULT 0,
            tax_rate REAL DEFAULT 0,
            party TEXT,
            remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    ''')

    # 发票表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            contract_id INTEGER,
            invoice_no TEXT,
            invoice_type TEXT,
            amount REAL DEFAULT 0,
            tax_rate REAL DEFAULT 0,
            tax_amount REAL DEFAULT 0,
            remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (contract_id) REFERENCES contracts(id)
        )
    ''')

    # 付款表（更新版）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            contract_id INTEGER,
            participant_id INTEGER,
            payment_no TEXT,
            amount REAL DEFAULT 0,
            payment_date DATE,
            payment_method TEXT,
            payer TEXT,
            remark TEXT,
            attachment TEXT,
            source TEXT DEFAULT 'manual',
            raw_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (contract_id) REFERENCES contracts(id),
            FOREIGN KEY (participant_id) REFERENCES participants(id)
        )
    ''')

    # 往来款表（更新版）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            participant_id INTEGER,
            trans_type TEXT,
            amount REAL DEFAULT 0,
            trans_date DATE,
            payment_method TEXT,
            payer TEXT,
            remark TEXT,
            attachment TEXT,
            source TEXT DEFAULT 'manual',
            raw_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (participant_id) REFERENCES participants(id)
        )
    ''')

    # 收支记录表（统一的交易表）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transaction_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            participant_id INTEGER,
            category_id INTEGER,
            trans_date DATE,
            amount REAL DEFAULT 0,
            trans_type TEXT,
            description TEXT,
            merchant TEXT,
            payment_method TEXT,
            attachment TEXT,
            source TEXT DEFAULT 'manual',
            raw_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (participant_id) REFERENCES participants(id)
        )
    ''')

    # 分类表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            type TEXT,
            keywords TEXT,
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 项目-分类白名单（每个项目启用的费用分类）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE,
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE,
            UNIQUE(project_id, category_id)
        )
    ''')

    # 投资记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS investments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            participant_id INTEGER,
            invest_date DATE,
            amount REAL DEFAULT 0,
            invest_type TEXT,
            payment_method TEXT,
            remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (participant_id) REFERENCES participants(id)
        )
    ''')

    # 分红记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS dividends (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            participant_id INTEGER,
            dividend_date DATE,
            amount REAL DEFAULT 0,
            period TEXT,
            payment_method TEXT,
            remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (participant_id) REFERENCES participants(id)
        )
    ''')

    # OCR草稿表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ocr_drafts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            image_path TEXT,
            draft_data TEXT,
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 操作日志表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT,
            action TEXT,
            detail TEXT,
            ip TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 采购单表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS purchase_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            contract_id INTEGER,
            purchase_no TEXT,
            supplier TEXT,
            order_date DATE,
            total_amount REAL DEFAULT 0,
            status TEXT DEFAULT '待审核',
            remark TEXT,
            attachment TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (contract_id) REFERENCES contracts(id)
        )
    ''')

    # 采购明细表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS purchase_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_id INTEGER,
            item_name TEXT,
            specification TEXT,
            unit TEXT,
            quantity REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            amount REAL DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (purchase_id) REFERENCES purchase_orders(id)
        )
    ''')

    # 销售出库单表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sales_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            contract_id INTEGER,
            order_no TEXT,
            customer_name TEXT,
            order_date DATE,
            delivery_date DATE,
            total_amount REAL DEFAULT 0,
            total_quantity REAL DEFAULT 0,
            status TEXT DEFAULT '待审核',
            remark TEXT,
            attachment TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (contract_id) REFERENCES contracts(id)
        )
    ''')

    # 销售出库明细表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sales_order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sales_order_id INTEGER,
            item_name TEXT,
            specification TEXT,
            unit TEXT,
            quantity REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            amount REAL DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (sales_order_id) REFERENCES sales_orders(id)
        )
    ''')

    # 运输记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transport_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            purchase_id INTEGER,
            batch_no TEXT,
            vehicle_no TEXT,
            driver_name TEXT,
            driver_phone TEXT,
            transport_date DATE,
            quantity REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            freight_amount REAL DEFAULT 0,
            remark TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (purchase_id) REFERENCES purchase_orders(id)
        )
    ''')

    # 运输-采购明细关联表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transport_purchase_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            transport_id INTEGER,
            purchase_item_id INTEGER,
            quantity REAL DEFAULT 0,
            FOREIGN KEY (transport_id) REFERENCES transport_records(id),
            FOREIGN KEY (purchase_item_id) REFERENCES purchase_items(id)
        )
    ''')

    # 销售-运输关联表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sales_item_transport (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sales_item_id INTEGER,
            transport_id INTEGER,
            quantity REAL DEFAULT 0,
            FOREIGN KEY (sales_item_id) REFERENCES sales_order_items(id),
            FOREIGN KEY (transport_id) REFERENCES transport_records(id)
        )
    ''')

    # 供应商表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            contact TEXT,
            phone TEXT,
            email TEXT,
            address TEXT,
            delivery_address TEXT,
            bank_name TEXT,
            bank_account TEXT,
            bank_code TEXT,
            tax_no TEXT,
            invoice_title TEXT,
            invoice_addr_phone TEXT,
            invoice_bank_account TEXT,
            tax_rate REAL DEFAULT 0,
            remark TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 客户表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            contact TEXT,
            phone TEXT,
            email TEXT,
            address TEXT,
            delivery_address TEXT,
            bank_name TEXT,
            bank_account TEXT,
            bank_code TEXT,
            tax_no TEXT,
            invoice_title TEXT,
            invoice_addr_phone TEXT,
            invoice_bank_account TEXT,
            tax_rate REAL DEFAULT 0,
            remark TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 付款类型表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payment_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 对账单表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reconciliations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            reconciliation_no TEXT,
            reconciliation_type TEXT,
            party_name TEXT,
            period_start DATE,
            period_end DATE,
            total_amount REAL DEFAULT 0,
            verified_amount REAL DEFAULT 0,
            difference REAL DEFAULT 0,
            status TEXT DEFAULT '待对账',
            remark TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    ''')

    # 销售回款表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sales_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            contract_id INTEGER,
            customer_name TEXT,
            payment_no TEXT,
            amount REAL DEFAULT 0,
            payment_date DATE,
            payment_method TEXT,
            remark TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id),
            FOREIGN KEY (contract_id) REFERENCES contracts(id)
        )
    ''')

    # 客户协同门户表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            phone TEXT,
            contact_name TEXT,
            company_name TEXT,
            status TEXT DEFAULT 'pending',
            balance REAL DEFAULT 0,
            total_recharge REAL DEFAULT 0,
            total_deduct REAL DEFAULT 0,
            alert_threshold REAL DEFAULT 10,
            last_alert_at TIMESTAMP,
            approved_by INTEGER,
            approved_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES customers(id),
            FOREIGN KEY (approved_by) REFERENCES users(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_recharges (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            amount REAL NOT NULL,
            payment_method TEXT,
            payment_no TEXT,
            remark TEXT,
            attachment TEXT,
            status TEXT DEFAULT 'pending',
            confirmed_by INTEGER,
            confirmed_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (client_id) REFERENCES client_accounts(id),
            FOREIGN KEY (confirmed_by) REFERENCES users(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_deductions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            sales_order_id INTEGER,
            sales_item_id INTEGER,
            amount REAL NOT NULL,
            quantity REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            item_name TEXT,
            deduct_date DATE,
            remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (client_id) REFERENCES client_accounts(id),
            FOREIGN KEY (sales_order_id) REFERENCES sales_orders(id),
            FOREIGN KEY (sales_item_id) REFERENCES sales_order_items(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            title TEXT,
            content TEXT,
            msg_type TEXT DEFAULT 'alert',
            is_read INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (client_id) REFERENCES client_accounts(id)
        )
    ''')

    # 创建默认管理员
    cursor.execute("SELECT COUNT(*) FROM users WHERE username='admin'")
    if cursor.fetchone()[0] == 0:
        hashed = hashlib.md5('admin123'.encode()).hexdigest()
        cursor.execute(
            "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
            ('admin', hashed, 'admin')
        )

    # 初始化默认分类
    init_categories(cursor)

    # 兼容旧数据库：添加缺失的列
    try:
        cursor.execute("ALTER TABLE sales_order_items ADD COLUMN sales_order_id INTEGER")
    except:
        pass  # 列已存在

    db.commit()
    db.close()


def init_categories(cursor):
    """初始化默认分类"""
    expense_categories = [
        ('原材料采购', 'expense', '原材料,采购,材料'),
        ('辅材采购', 'expense', '辅材,辅助材料'),
        ('低值易耗品', 'expense', '低值易耗'),
        ('运输装卸', 'expense', '运输,装卸,物流'),
        ('仓储保管费', 'expense', '仓储,保管,仓库'),
        ('机械设备', 'expense', '机械,设备'),
        ('设备租赁费', 'expense', '租赁,设备租赁'),
        ('工程施工', 'expense', '施工,工程'),
        ('劳务费', 'expense', '劳务,人工'),
        ('外协加工费', 'expense', '外协,加工'),
        ('检测检验费', 'expense', '检测,检验'),
        ('设计制图费', 'expense', '设计,制图'),
        ('办公用品', 'expense', '办公用品,文具'),
        ('办公费', 'expense', '办公'),
        ('通讯费', 'expense', '通讯,电话,手机'),
        ('水电费', 'expense', '水电,水费,电费'),
        ('物业管理费', 'expense', '物业'),
        ('清洁绿化费', 'expense', '清洁,绿化'),
        ('维修费', 'expense', '维修,修理'),
        ('软件服务费', 'expense', '软件,服务'),
        ('快递物流费', 'expense', '快递,物流'),
        ('安全生产费', 'expense', '安全'),
        ('差旅费', 'expense', '差旅'),
        ('交通费', 'expense', '交通,打车,地铁'),
        ('住宿费', 'expense', '住宿,酒店'),
        ('餐饮费', 'expense', '餐饮,餐费,吃饭'),
        ('招待费', 'expense', '招待'),
        ('会议费', 'expense', '会议'),
        ('培训费', 'expense', '培训'),
        ('礼品费', 'expense', '礼品,礼物'),
        ('广告宣传费', 'expense', '广告,宣传'),
        ('市场推广费', 'expense', '市场,推广'),
        ('业务服务费', 'expense', '业务,服务'),
        ('咨询服务费', 'expense', '咨询'),
        ('法务费', 'expense', '法务,律师'),
        ('审计费', 'expense', '审计'),
        ('中介服务费', 'expense', '中介'),
        ('银行手续费', 'expense', '手续费,银行'),
        ('税费支出', 'expense', '税费,税金'),
        ('工资薪酬', 'expense', '工资,薪酬,薪水'),
        ('福利费', 'expense', '福利'),
        ('社保公积金', 'expense', '社保,公积金'),
        ('保险费', 'expense', '保险'),
        ('招聘费', 'expense', '招聘'),
        ('燃油费', 'expense', '燃油,汽油,柴油'),
        ('过路过桥费', 'expense', '过路,过桥,高速'),
        ('停车费', 'expense', '停车'),
        ('折旧摊销', 'expense', '折旧,摊销'),
        ('其他支出', 'expense', '其他'),
    ]

    income_categories = [
        ('项目回款', 'income', '回款,收款'),
        ('项目投资', 'income', '投资'),
        ('其他收入', 'income', '其他'),
    ]

    cursor.execute("SELECT COUNT(*) FROM categories")
    if cursor.fetchone()[0] == 0:
        for name, type_, keywords in expense_categories:
            cursor.execute(
                "INSERT INTO categories (name, type, keywords) VALUES (?, ?, ?)",
                (name, type_, keywords)
            )
        for name, type_, keywords in income_categories:
            cursor.execute(
                "INSERT INTO categories (name, type, keywords) VALUES (?, ?, ?)",
                (name, type_, keywords)
            )


def add_log(user_id, username, action, detail, ip=''):
    """记录操作日志"""
    try:
        db = get_db()
        db.execute(
            "INSERT INTO logs (user_id, username, action, detail, ip) VALUES (?, ?, ?, ?, ?)",
            (user_id, username, action, detail, ip)
        )
        db.commit()
    except Exception:
        pass


# ==================== 登录装饰器 ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('权限不足，仅管理员可操作', 'danger')
            if session.get('role') == ROLE_CLIENT_COLLAB:
                return redirect(url_for('admin_client_dashboard'))
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== OCR 识别工具 ====================

def ocr_recognize(image_path):
    """
    OCR 识别图片中的交易信息
    返回识别出的多条交易记录
    优先 RapidOCR（项目自带 ocr_utils），回退 PaddleOCR / EasyOCR
    """
    text = ""
    engine_used = None

    # 方法1: RapidOCR（推荐，依赖少）
    try:
        from ocr_utils import ocr_image_to_lines
        lines = ocr_image_to_lines(image_path)
        if lines:
            text = '\n'.join(lines)
            engine_used = 'RapidOCR'
            print(f"[RapidOCR] 识别完成，{len(lines)} 行文字")
    except ImportError as e:
        print(f"[RapidOCR] 未安装: {e}")
    except Exception as e:
        print(f"[RapidOCR] 错误: {e}")

    # 方法2: PaddleOCR
    if not text or len(text.strip()) < 5:
        try:
            from paddleocr import PaddleOCR
            os.environ['KMP_DUPLICATE_LIB_OK'] = 'TRUE'
            ocr = PaddleOCR(use_angle_cls=True, lang='ch', use_gpu=False, show_log=False)
            result = ocr.ocr(image_path, cls=True)
            if result and result[0]:
                lines = []
                for line in result[0]:
                    if line and len(line) >= 2 and line[1][1] > 0.5:
                        lines.append(line[1][0])
                if lines:
                    text = '\n'.join(lines)
                    engine_used = 'PaddleOCR'
                    print(f"[PaddleOCR] 识别完成，{len(lines)} 行文字")
        except ImportError:
            print("[PaddleOCR] 未安装")
        except Exception as e:
            print(f"[PaddleOCR] 错误: {e}")

    # 方法3: EasyOCR
    if not text or len(text.strip()) < 5:
        try:
            import easyocr
            reader = easyocr.Reader(['ch_sim', 'en'], gpu=False, verbose=False)
            result = reader.readtext(image_path, detail=1)
            sorted_result = sorted(
                result, key=lambda x: (round(x[0][0][1] / 20) * 20, x[0][0][0])
            )
            lines = [item[1] for item in sorted_result if item[2] > 0.3]
            if lines:
                text = '\n'.join(lines)
                engine_used = 'EasyOCR'
                print(f"[EasyOCR] 识别完成，{len(lines)} 行文字")
        except ImportError:
            print("[EasyOCR] 未安装")
        except Exception as e:
            print(f"[EasyOCR] 错误: {e}")

    if not text or len(text.strip()) < 5:
        install_hint = (
            "无法识别文字。请安装 OCR 组件：pip install rapidocr-onnxruntime\n"
            "（或 pip install paddleocr / easyocr）"
        )
        try:
            from PIL import Image
            img = Image.open(image_path)
            text = f"[图片尺寸: {img.size}]\n{install_hint}"
        except Exception as e:
            text = f"[无法读取图片: {e}]\n{install_hint}"
    transactions = parse_ocr_text(text)
    return transactions, text


def _ocr_parse_date(line):
    """从一行文本解析日期 YYYY-MM-DD"""
    m = re.search(
        r'(\d{4})\s*[-/年\.]\s*(\d{1,2})\s*[-/月\.]\s*(\d{1,2})',
        line,
    )
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    if '今天' in line:
        return datetime.now().strftime('%Y-%m-%d')
    if '昨天' in line:
        return (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    return None


def _ocr_parse_amount(line):
    """从一行解析金额，避免把日期里的年份当成金额"""
    if re.search(r'\d{4}\s*[-/年]', line) and not re.search(r'\d+\.\d{2}', line):
        # 纯日期行不在此行取金额（除非带小数金额）
        pass
    patterns = [
        r'[¥￥]\s*(-?\d{1,3}(?:,\d{3})*\.\d{2})',
        r'(-?\d{1,3}(?:,\d{3})*\.\d{2})',
        r'[¥￥]\s*(-?\d+(?:\.\d+)?)',
    ]
    for pat in patterns:
        m = re.search(pat, line)
        if m:
            try:
                val = float(m.group(1).replace(',', ''))
                if 0 < abs(val) < 50_000_000:
                    return abs(val)
            except ValueError:
                continue
    return 0.0


def _ocr_line_meta(line):
    """单行提取商户、支付方式、收支类型"""
    merchant = ''
    for pat in (
        r'(?:商户|对方|收款方|付款方)[:：\s]*(.+)',
        r'((?:微信|支付宝|云闪付|美团|饿了么|滴滴|京东|淘宝|拼多多|抖音)[^\d]{0,20})',
    ):
        m = re.search(pat, line)
        if m and len(m.group(1).strip()) > 1:
            merchant = m.group(1).strip()[:80]
            break
    payment_method = ''
    for pm in ('微信支付', '支付宝', '云闪付', '银行卡', '花呗', '信用卡', '零钱'):
        if pm in line:
            payment_method = pm
            break
    trans_type = 'income' if any(k in line for k in ('收入', '收款', '入账', '退款')) else 'expense'
    return merchant, payment_method, trans_type


def parse_ocr_text(text):
    """解析 OCR 文本，提取多条交易记录"""
    if not text or '请安装' in text or '无法识别文字' in text or text.startswith('[图片尺寸'):
        return []
    lines = [ln.strip() for ln in text.replace('\r', '').split('\n') if ln.strip()]
    if not lines:
        return []

    try:
        print(f"[OCR原始文本] {len(lines)} 行:\n" + '\n'.join(lines[:20]))
    except UnicodeEncodeError:
        print(f"[OCR原始文本] {len(lines)} 行")

    found_records = []
    pending_date = None
    pending_meta = ('', '', 'expense')

    for i, line in enumerate(lines):
        date_str = _ocr_parse_date(line)
        amount = _ocr_parse_amount(line)
        merchant, payment_method, trans_type = _ocr_line_meta(line)

        if date_str:
            pending_date = date_str
        if merchant:
            pending_meta = (merchant, payment_method, trans_type)

        if amount > 0:
            use_date = date_str or pending_date or datetime.now().strftime('%Y-%m-%d')
            m, pm, tt = pending_meta
            if merchant:
                m, pm, tt = merchant, payment_method or pm, trans_type
            found_records.append({
                'date': use_date[:10] if len(use_date) >= 10 else use_date,
                'amount': amount,
                'trans_type': tt,
                'merchant': m,
                'payment_method': pm,
                'description': m or line[:80],
                'raw_text': line[:200],
            })
            if date_str:
                pending_date = date_str

    # 去重
    seen = set()
    transactions = []
    for rec in found_records:
        key = f"{rec['date']}_{rec['amount']}_{rec.get('merchant', '')}"
        if key not in seen:
            seen.add(key)
            transactions.append(rec)

    print(f"[识别结果] 找到 {len(transactions)} 条")
    return transactions


def match_category(description, trans_type='expense'):
    """根据描述匹配分类"""
    db = get_db()
    categories = db.execute(
        "SELECT * FROM categories WHERE type=?", (trans_type,)
    ).fetchall()

    if not description:
        return None

    description = description.lower()
    for cat in categories:
        if cat['keywords']:
            keywords = cat['keywords'].split(',')
            for kw in keywords:
                if kw.lower() in description:
                    return cat['id']
    return None


def match_participant(text, project_id=None):
    """根据文本匹配参与人"""
    db = get_db()

    # 如果指定了项目，优先在项目参与人中匹配
    if project_id:
        participants = db.execute("""
            SELECT p.* FROM participants p
            JOIN project_participants pp ON p.id = pp.participant_id
            WHERE pp.project_id = ?
        """, (project_id,)).fetchall()
    else:
        participants = db.execute("SELECT * FROM participants").fetchall()

    for p in participants:
        # 匹配姓名
        if p['name'] and p['name'] in text:
            return p['id']
        # 匹配手机号
        if p['phone'] and p['phone'] in text:
            return p['id']

    # 匹配账户尾号
    accounts = db.execute("SELECT * FROM participant_accounts").fetchall()
    for acc in accounts:
        if acc['account_tail'] and acc['account_tail'] in text:
            return acc['participant_id']
        if acc['keywords']:
            for kw in acc['keywords'].split(','):
                if kw in text:
                    return acc['participant_id']

    return None


# ==================== Excel 导入工具 ====================

def parse_excel(file_path):
    """解析 Excel 文件"""
    try:
        import pandas as pd
        df = pd.read_excel(file_path)
        return df.to_dict('records')
    except Exception as e:
        return {'error': str(e)}


def import_from_excel(records, project_name=None, user_id=None):
    """从 Excel 记录导入数据"""
    db = get_db()
    results = {
        'imported': 0,
        'updated': 0,
        'skipped': 0,
        'project_created': 0,
        'participant_created': 0,
        'errors': []
    }

    for row in records:
        try:
            # 解析项目名称
            proj_name = project_name or row.get('项目') or row.get('project') or row.get('项目名称')
            if not proj_name:
                results['skipped'] += 1
                continue

            # 获取或创建项目
            project = db.execute(
                "SELECT * FROM projects WHERE name=?", (proj_name,)
            ).fetchone()

            if not project:
                db.execute(
                    "INSERT INTO projects (name, status, user_id) VALUES (?, '进行中', ?)",
                    (proj_name, user_id)
                )
                db.commit()
                project = db.execute(
                    "SELECT * FROM projects WHERE name=?", (proj_name,)
                ).fetchone()
                results['project_created'] += 1

            project_id = project['id']

            # 解析参与人
            participant_name = row.get('参与人') or row.get('participant') or row.get('姓名')
            participant_id = None
            if participant_name:
                participant = db.execute(
                    "SELECT * FROM participants WHERE name=?", (participant_name,)
                ).fetchone()
                if not participant:
                    db.execute(
                        "INSERT INTO participants (name, role) VALUES (?, 'member')",
                        (participant_name,)
                    )
                    db.commit()
                    participant = db.execute(
                        "SELECT * FROM participants WHERE name=?", (participant_name,)
                    ).fetchone()
                    results['participant_created'] += 1

                    # 绑定到项目
                    db.execute(
                        "INSERT OR IGNORE INTO project_participants (project_id, participant_id) VALUES (?, ?)",
                        (project_id, participant['id'])
                    )

                participant_id = participant['id']

            # 解析日期
            date_str = row.get('日期') or row.get('date') or row.get('时间')
            if isinstance(date_str, datetime):
                trans_date = date_str.strftime('%Y-%m-%d')
            else:
                trans_date = str(date_str)[:10] if date_str else datetime.now().strftime('%Y-%m-%d')

            # 解析金额
            amount = row.get('金额') or row.get('amount') or 0
            try:
                amount = float(amount)
            except:
                amount = 0

            # 解析类型
            type_str = str(row.get('类型') or row.get('type') or '').lower()
            if '收入' in type_str or 'income' in type_str or '回款' in type_str:
                trans_type = 'income'
            else:
                trans_type = 'expense'

            # 解析说明和商户
            description = row.get('说明') or row.get('description') or row.get('备注') or ''
            merchant = row.get('商户') or row.get('merchant') or row.get('对方') or ''

            # 解析支付方式
            payment_method = row.get('支付方式') or row.get('payment_method') or ''

            # 匹配分类
            category_id = match_category(description or merchant, trans_type)

            # 检查是否已存在（根据日期、金额、说明去重）
            existing = db.execute("""
                SELECT * FROM transaction_records
                WHERE project_id=? AND trans_date=? AND amount=? AND description=?
            """, (project_id, trans_date, amount, description)).fetchone()

            raw_data = json.dumps(row, ensure_ascii=False, default=str)

            if existing:
                # 更新现有记录
                db.execute("""
                    UPDATE transaction_records SET
                        participant_id=?, category_id=?, trans_type=?,
                        merchant=?, payment_method=?, raw_data=?, source='excel'
                    WHERE id=?
                """, (participant_id, category_id, trans_type, merchant, payment_method, raw_data, existing['id']))
                results['updated'] += 1
            else:
                # 插入新记录
                db.execute("""
                    INSERT INTO transaction_records
                    (project_id, participant_id, category_id, trans_date, amount, trans_type,
                     description, merchant, payment_method, source, raw_data)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'excel', ?)
                """, (project_id, participant_id, category_id, trans_date, amount, trans_type,
                      description, merchant, payment_method, raw_data))
                results['imported'] += 1

            db.commit()

        except Exception as e:
            results['errors'].append(str(e))
            results['skipped'] += 1

    return results


# ==================== 路由：登录/登出 ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            flash('请输入用户名和密码', 'warning')
            return render_template('login.html')

        hashed = hashlib.md5(password.encode()).hexdigest()
        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE username=? AND password=?",
            (username, hashed)
        ).fetchone()

        if user:
            status = user['status'] if 'status' in user.keys() and user['status'] else 'active'
            if status in ('pending', 'disabled'):
                flash('账号未启用或待审批，请联系管理员', 'danger')
                return render_template('login.html')
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            try:
                db.execute(
                    "UPDATE users SET last_login=? WHERE id=?",
                    (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']),
                )
                db.commit()
            except sqlite3.OperationalError:
                pass
            add_log(user['id'], user['username'], '登录', '用户登录系统', request.remote_addr)
            return redirect(login_redirect_for_role(user['role']))
        else:
            flash('用户名或密码错误', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    user_id = session.get('user_id')
    username = session.get('username')
    if user_id:
        add_log(user_id, username, '登出', '用户登出系统', request.remote_addr)
    session.clear()
    return redirect(url_for('login'))


# ==================== 路由：看板/仪表盘 ====================

@app.route('/')
@login_required
def dashboard():
    db = get_db()
    user_id = session['user_id']
    role = session['role']

    # 根据角色筛选项目
    if role in ('admin', 'finance'):
        projects = db.execute("SELECT * FROM projects ORDER BY created_at DESC").fetchall()
    else:
        projects = db.execute(
            "SELECT * FROM projects WHERE user_id=? ORDER BY created_at DESC",
            (user_id,)
        ).fetchall()

    project_ids = [p['id'] for p in projects]

    # 统计数据
    total_projects = len(projects)

    # 从交易记录统计
    total_income = 0
    total_expense = 0
    if project_ids:
        placeholders = ','.join(['?'] * len(project_ids))
        income_rows = db.execute(
            f"SELECT COALESCE(SUM(amount),0) as total FROM transaction_records WHERE project_id IN ({placeholders}) AND trans_type='income'",
            project_ids
        ).fetchone()
        total_income = income_rows['total'] if income_rows else 0

        expense_rows = db.execute(
            f"SELECT COALESCE(SUM(amount),0) as total FROM transaction_records WHERE project_id IN ({placeholders}) AND trans_type='expense'",
            project_ids
        ).fetchone()
        total_expense = expense_rows['total'] if expense_rows else 0

    profit = total_income - total_expense
    profit_rate = (profit / total_income * 100) if total_income > 0 else 0

    # 最近交易记录
    recent_transactions = db.execute(
        """SELECT t.*, p.name as project_name, cat.name as category_name
           FROM transaction_records t
           LEFT JOIN projects p ON t.project_id = p.id
           LEFT JOIN categories cat ON t.category_id = cat.id
           ORDER BY t.created_at DESC LIMIT 5"""
    ).fetchall()

    # 最近操作日志
    logs = db.execute(
        "SELECT * FROM logs ORDER BY created_at DESC LIMIT 10"
    ).fetchall()

    return render_template('dashboard.html',
                           projects=projects,
                           total_projects=total_projects,
                           total_income=total_income,
                           total_expense=total_expense,
                           profit=profit,
                           profit_rate=profit_rate,
                           recent_transactions=recent_transactions,
                           logs=logs)


# ==================== 路由：项目管理 ====================

@app.route('/projects')
@login_required
def project_list():
    db = get_db()
    user_id = session['user_id']
    role = session['role']

    if role in ('admin', 'finance'):
        projects = db.execute("SELECT * FROM projects ORDER BY created_at DESC").fetchall()
    else:
        projects = db.execute(
            "SELECT * FROM projects WHERE user_id=? ORDER BY created_at DESC",
            (user_id,)
        ).fetchall()

    # 为每个项目计算收支
    project_data = []
    for p in projects:
        income = db.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM transaction_records WHERE project_id=? AND trans_type='income'",
            (p['id'],)
        ).fetchone()['total']
        expense = db.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM transaction_records WHERE project_id=? AND trans_type='expense'",
            (p['id'],)
        ).fetchone()['total']

        # 计算投资
        investment = db.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM investments WHERE project_id=?",
            (p['id'],)
        ).fetchone()['total']

        # 计算分红
        dividend = db.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM dividends WHERE project_id=?",
            (p['id'],)
        ).fetchone()['total']

        project_data.append({
            'project': p,
            'income': income,
            'expense': expense,
            'investment': investment,
            'dividend': dividend,
            'balance': investment + income - expense - dividend
        })

    return render_template('project_list.html', project_data=project_data)


@app.route('/project/add', methods=['GET', 'POST'])
@login_required
def project_add():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        status = request.form.get('status', '进行中')
        budget = float(request.form.get('budget', 0) or 0)
        start_date = request.form.get('start_date', '')
        end_date = request.form.get('end_date', '')

        if not name:
            flash('项目名称不能为空', 'warning')
            return render_template('project_form.html')

        db = get_db()
        db.execute(
            """INSERT INTO projects (name, description, status, budget, start_date, end_date, user_id)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (name, description, status, budget, start_date, end_date, session['user_id'])
        )
        db.commit()
        add_log(session['user_id'], session['username'], '新增项目', f'新增项目: {name}', request.remote_addr)
        flash('项目创建成功', 'success')
        return redirect(url_for('project_list'))

    return render_template('project_form.html')


@app.route('/project/delete/<int:pid>')
@login_required
def project_delete(pid):
    db = get_db()
    role = session['role']
    if role != 'admin':
        flash('仅管理员可删除项目', 'danger')
        return redirect(url_for('project_list'))

    project = db.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    if project:
        # 删除关联数据
        db.execute("DELETE FROM transaction_records WHERE project_id=?", (pid,))
        db.execute("DELETE FROM transactions WHERE project_id=?", (pid,))
        db.execute("DELETE FROM payments WHERE project_id=?", (pid,))
        db.execute("DELETE FROM invoices WHERE project_id=?", (pid,))
        db.execute("DELETE FROM contracts WHERE project_id=?", (pid,))
        db.execute("DELETE FROM investments WHERE project_id=?", (pid,))
        db.execute("DELETE FROM dividends WHERE project_id=?", (pid,))
        db.execute("DELETE FROM project_participants WHERE project_id=?", (pid,))
        db.execute("DELETE FROM projects WHERE id=?", (pid,))
        db.commit()
        add_log(session['user_id'], session['username'], '删除项目', f'删除项目: {project["name"]}', request.remote_addr)
        flash('项目已删除', 'success')
    return redirect(url_for('project_list'))


# ==================== 路由：项目详情 ====================

def recalc_investment_ratios(db, pid):
    """根据各参与人在本项目的实际投资金额，自动计算并写回投资比例（占项目总投资的百分比）。"""
    total = db.execute(
        "SELECT COALESCE(SUM(amount), 0) FROM investments WHERE project_id=?", (pid,)
    ).fetchone()[0] or 0
    rows = db.execute(
        "SELECT participant_id FROM project_participants WHERE project_id=?", (pid,)
    ).fetchall()
    for r in rows:
        part_sum = db.execute(
            "SELECT COALESCE(SUM(amount), 0) FROM investments WHERE project_id=? AND participant_id=?",
            (pid, r['participant_id'])
        ).fetchone()[0] or 0
        ratio = round(part_sum / total * 100, 2) if total > 0 else 0
        db.execute(
            "UPDATE project_participants SET investment_ratio=? WHERE project_id=? AND participant_id=?",
            (ratio, pid, r['participant_id'])
        )
    db.commit()


@app.route('/project/<int:pid>')
@login_required
def project_detail(pid):
    db = get_db()
    project = db.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    if not project:
        flash('项目不存在', 'danger')
        return redirect(url_for('project_list'))

    # 投资比例按实际投资金额自动重算（自愈，保证展示与下游分红计算一致）
    recalc_investment_ratios(db, pid)

    # 参与人列表
    participants = db.execute("""
        SELECT p.*, pp.project_role, pp.investment_ratio, pp.dividend_ratio
        FROM participants p
        JOIN project_participants pp ON p.id = pp.participant_id
        WHERE pp.project_id=?
    """, (pid,)).fetchall()

    # 投资记录
    investments = db.execute("""
        SELECT i.*, p.name as participant_name
        FROM investments i
        LEFT JOIN participants p ON i.participant_id = p.id
        WHERE i.project_id=? ORDER BY i.invest_date DESC
    """, (pid,)).fetchall()

    # 分红记录
    dividends = db.execute("""
        SELECT d.*, p.name as participant_name
        FROM dividends d
        LEFT JOIN participants p ON d.participant_id = p.id
        WHERE d.project_id=? ORDER BY d.dividend_date DESC
    """, (pid,)).fetchall()

    # 交易记录
    transactions = db.execute("""
        SELECT t.*, p.name as participant_name, cat.name as category_name
        FROM transaction_records t
        LEFT JOIN participants p ON t.participant_id = p.id
        LEFT JOIN categories cat ON t.category_id = cat.id
        WHERE t.project_id=? ORDER BY t.trans_date DESC
    """, (pid,)).fetchall()

    # 付款记录
    payments = db.execute("""
        SELECT pay.*, p.name as participant_name, c.contract_name
        FROM payments pay
        LEFT JOIN participants p ON pay.participant_id = p.id
        LEFT JOIN contracts c ON pay.contract_id = c.id
        WHERE pay.project_id=? ORDER BY pay.payment_date DESC
    """, (pid,)).fetchall()

    # 往来款记录
    fund_trans = db.execute("""
        SELECT ft.*, p.name as participant_name
        FROM transactions ft
        LEFT JOIN participants p ON ft.participant_id = p.id
        WHERE ft.project_id=? ORDER BY ft.trans_date DESC
    """, (pid,)).fetchall()

    # 统计
    total_income = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transaction_records WHERE project_id=? AND trans_type='income'",
        (pid,)
    ).fetchone()[0]
    total_expense = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transaction_records WHERE project_id=? AND trans_type='expense'",
        (pid,)
    ).fetchone()[0]
    total_investment = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM investments WHERE project_id=?",
        (pid,)
    ).fetchone()[0]
    total_dividend = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM dividends WHERE project_id=?",
        (pid,)
    ).fetchone()[0]

    balance = total_investment + total_income - total_expense - total_dividend

    available_participants = db.execute("""
        SELECT * FROM participants p
        WHERE p.id NOT IN (
            SELECT participant_id FROM project_participants WHERE project_id=?
        )
        ORDER BY p.name
    """, (pid,)).fetchall()

    ensure_project_categories_table(db)
    cat_filter_sql, cat_filter_params = category_ids_sql_filter(db, pid, "c.id")
    category_summary = db.execute(f"""
        SELECT c.id, c.name, c.type,
               COALESCE(SUM(t.amount), 0) as total,
               COUNT(t.id) as cnt
        FROM transaction_records t
        JOIN categories c ON t.category_id = c.id
        WHERE t.project_id=?{cat_filter_sql}
        GROUP BY c.id
        ORDER BY c.type, total DESC
    """, [pid] + cat_filter_params).fetchall()

    categories = get_project_categories_list(db, pid)
    all_categories = fetch_leaf_categories(db)
    enabled_category_ids = get_project_enabled_category_ids(db, pid)
    uses_category_whitelist = project_uses_category_whitelist(db, pid)

    return render_template('project_detail.html',
                           project=project,
                           participants=participants,
                           available_participants=available_participants,
                           investments=investments,
                           dividends=dividends,
                           transactions=transactions,
                           payments=payments,
                           fund_trans=fund_trans,
                           total_income=total_income,
                           total_expense=total_expense,
                           total_investment=total_investment,
                           total_dividend=total_dividend,
                           balance=balance,
                           category_summary=category_summary,
                           categories=categories,
                           all_categories=all_categories,
                           enabled_category_ids=enabled_category_ids,
                           uses_category_whitelist=uses_category_whitelist,
                           now=datetime.now())


@app.route('/project/<int:pid>/categories', methods=['POST'])
@login_required
def project_categories_save(pid):
    db = get_db()
    project = db.execute("SELECT id FROM projects WHERE id=?", (pid,)).fetchone()
    if not project:
        flash('项目不存在', 'danger')
        return redirect(url_for('project_list'))
    ensure_project_categories_table(db)
    category_ids = request.form.getlist('category_ids')
    set_project_categories(db, pid, category_ids)
    db.commit()
    add_log(session['user_id'], session['username'], '配置项目分类',
            f'项目{pid}启用{len(category_ids)}个费用分类', request.remote_addr)
    flash(f'已保存，本项目启用 {len(category_ids)} 个费用分类', 'success')
    return redirect(url_for('project_detail', pid=pid) + '#tab-categories')


# ==================== 路由：参与人管理 ====================

@app.route('/participants')
@login_required
def participant_list():
    db = get_db()
    participants = db.execute("""
        SELECT p.*,
               (SELECT COUNT(*) FROM project_participants WHERE participant_id=p.id) as project_count,
               (SELECT COALESCE(SUM(amount),0) FROM investments WHERE participant_id=p.id) as total_investment,
               (SELECT COALESCE(SUM(amount),0) FROM transaction_records WHERE participant_id=p.id AND trans_type='expense') as total_expense,
               (SELECT COALESCE(SUM(amount),0) FROM dividends WHERE participant_id=p.id) as total_dividend
        FROM participants p ORDER BY p.created_at DESC
    """).fetchall()
    return render_template('participant_list.html', participants=participants)


@app.route('/participant/add', methods=['GET', 'POST'])
@login_required
def participant_add():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        phone = request.form.get('phone', '').strip()
        role = request.form.get('role', 'member')
        remark = request.form.get('remark', '').strip()

        if not name:
            flash('姓名不能为空', 'warning')
            return render_template('participant_form.html')

        db = get_db()
        db.execute(
            "INSERT INTO participants (name, phone, role, remark) VALUES (?, ?, ?, ?)",
            (name, phone, role, remark)
        )
        db.commit()
        add_log(session['user_id'], session['username'], '新增参与人', f'新增参与人: {name}', request.remote_addr)
        flash('参与人创建成功', 'success')
        return redirect(url_for('participant_list'))

    return render_template('participant_form.html')


@app.route('/project/<int:pid>/participant/add', methods=['POST'])
@login_required
def project_participant_add(pid):
    participant_id = request.form.get('participant_id')
    project_role = request.form.get('project_role', 'member')
    investment_ratio = float(request.form.get('investment_ratio', 0) or 0)
    dividend_ratio = float(request.form.get('dividend_ratio', 0) or 0)

    if not participant_id:
        flash('请选择参与人', 'warning')
        return redirect(url_for('project_detail', pid=pid))

    db = get_db()
    db.execute(
        """INSERT OR REPLACE INTO project_participants
           (project_id, participant_id, project_role, investment_ratio, dividend_ratio)
           VALUES (?, ?, ?, ?, ?)""",
        (pid, participant_id, project_role, investment_ratio, dividend_ratio)
    )
    db.commit()
    flash('参与人已添加到项目', 'success')
    return redirect(url_for('project_detail', pid=pid))


# ==================== 路由：投资记录 ====================

@app.route('/project/<int:pid>/investment/add', methods=['POST'])
@login_required
def investment_add(pid):
    participant_id = request.form.get('participant_id')
    invest_date = request.form.get('invest_date')
    amount = float(request.form.get('amount', 0) or 0)
    invest_type = request.form.get('invest_type', '启动资金')
    payment_method = request.form.get('payment_method', '')
    remark = request.form.get('remark', '')

    db = get_db()
    db.execute(
        """INSERT INTO investments (project_id, participant_id, invest_date, amount, invest_type, payment_method, remark)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (pid, participant_id, invest_date, amount, invest_type, payment_method, remark)
    )
    db.commit()
    recalc_investment_ratios(db, pid)
    add_log(session['user_id'], session['username'], '新增投资',
            f'项目{pid}新增{invest_type}: {amount}元', request.remote_addr)
    flash('投资记录添加成功', 'success')
    return redirect(url_for('project_detail', pid=pid))


# ==================== 路由：分红记录 ====================

@app.route('/project/<int:pid>/dividend/add', methods=['POST'])
@login_required
def dividend_add(pid):
    participant_id = request.form.get('participant_id')
    dividend_date = request.form.get('dividend_date')
    amount = float(request.form.get('amount', 0) or 0)
    period = request.form.get('period', '')
    payment_method = request.form.get('payment_method', '')
    remark = request.form.get('remark', '')

    db = get_db()
    db.execute(
        """INSERT INTO dividends (project_id, participant_id, dividend_date, amount, period, payment_method, remark)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (pid, participant_id, dividend_date, amount, period, payment_method, remark)
    )
    db.commit()
    add_log(session['user_id'], session['username'], '新增分红',
            f'项目{pid}新增分红: {amount}元', request.remote_addr)
    flash('分红记录添加成功', 'success')
    return redirect(url_for('project_detail', pid=pid))


# ==================== 路由：交易记录（收支） ====================

@app.route('/transactions')
@login_required
def transaction_records():
    db = get_db()

    # 筛选条件
    project_id = request.args.get('project_id', '')
    trans_type = request.args.get('type', '')

    where_sql = "WHERE 1=1"
    params = []

    if project_id:
        where_sql += " AND t.project_id=?"
        params.append(project_id)
    if trans_type:
        where_sql += " AND t.trans_type=?"
        params.append(trans_type)

    query = f"""
        SELECT t.*, p.name as project_name, par.name as participant_name, cat.name as category_name
        FROM transaction_records t
        LEFT JOIN projects p ON t.project_id = p.id
        LEFT JOIN participants par ON t.participant_id = par.id
        LEFT JOIN categories cat ON t.category_id = cat.id
        {where_sql}
        ORDER BY t.trans_date DESC
    """

    transactions = db.execute(query, params).fetchall()
    projects = db.execute("SELECT * FROM projects ORDER BY name").fetchall()
    ensure_project_categories_table(db)
    cat_filter_sql, cat_filter_params = "", []
    if project_id:
        categories = get_project_categories_list(db, int(project_id))
        cat_filter_sql, cat_filter_params = category_ids_sql_filter(
            db, int(project_id), "c.id"
        )
    else:
        categories = fetch_leaf_categories(db)

    category_summary = db.execute(f"""
        SELECT c.name, c.type,
               COALESCE(SUM(t.amount), 0) as total,
               COUNT(t.id) as cnt
        FROM transaction_records t
        JOIN categories c ON t.category_id = c.id
        {where_sql}{cat_filter_sql}
        GROUP BY c.id
        ORDER BY c.type, total DESC
    """, params + cat_filter_params).fetchall()

    total_income = sum(t['amount'] for t in transactions if t['trans_type'] == 'income')
    total_expense = sum(t['amount'] for t in transactions if t['trans_type'] == 'expense')

    return render_template('transaction_records.html',
                           transactions=transactions,
                           projects=projects,
                           categories=categories,
                           category_summary=category_summary,
                           project_id=project_id,
                           trans_type=trans_type,
                           total_income=total_income,
                           total_expense=total_expense)


@app.route('/transaction/add', methods=['GET', 'POST'])
@login_required
def transaction_add():
    db = get_db()
    projects = db.execute("SELECT * FROM projects ORDER BY name").fetchall()
    participants = db.execute("SELECT * FROM participants ORDER BY name").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY type, name").fetchall()

    if request.method == 'POST':
        project_id = request.form.get('project_id')
        participant_id = request.form.get('participant_id') or None
        category_id = request.form.get('category_id') or None
        trans_date = request.form.get('trans_date')
        amount = float(request.form.get('amount', 0) or 0)
        trans_type = request.form.get('trans_type', 'expense')
        description = request.form.get('description', '')
        merchant = request.form.get('merchant', '')
        payment_method = request.form.get('payment_method', '')

        ensure_project_categories_table(db)
        if project_id and category_id and not validate_project_category(
                db, project_id, category_id):
            flash('所选费用分类未在本项目启用，请先在项目详情中配置费用分类', 'danger')
            return redirect(url_for('transaction_add'))

        db.execute(
            """INSERT INTO transaction_records
               (project_id, participant_id, category_id, trans_date, amount, trans_type,
                description, merchant, payment_method)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (project_id, participant_id, category_id, trans_date, amount, trans_type,
             description, merchant, payment_method)
        )
        db.commit()
        add_log(session['user_id'], session['username'], '新增交易',
                f'新增{"收入" if trans_type=="income" else "支出"}: {amount}元', request.remote_addr)
        flash('交易记录添加成功', 'success')
        return redirect(url_for('transaction_records'))

    return render_template('transaction_form.html',
                           projects=projects,
                           participants=participants,
                           categories=categories,
                           now=datetime.now())


# ==================== 路由：付款管理（更新版） ====================

@app.route('/project/<int:pid>/payment/add', methods=['GET', 'POST'])
@login_required
def payment_add(pid):
    db = get_db()
    contracts = db.execute(
        "SELECT * FROM contracts WHERE project_id=? AND contract_type='支出合同'", (pid,)
    ).fetchall()
    participants = db.execute("""
        SELECT p.* FROM participants p
        JOIN project_participants pp ON p.id = pp.participant_id
        WHERE pp.project_id=?
    """, (pid,)).fetchall()

    if request.method == 'POST':
        contract_id = request.form.get('contract_id') or None
        participant_id = request.form.get('participant_id') or None
        payment_no = request.form.get('payment_no', '').strip()
        amount = float(request.form.get('amount', 0) or 0)
        payment_date = request.form.get('payment_date', '')
        payment_method = request.form.get('payment_method', '')
        payer = request.form.get('payer', '')
        remark = request.form.get('remark', '')

        # 处理附件上传
        attachment = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename:
                filename = f"payment_{uuid.uuid4().hex[:8]}_{file.filename}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                attachment = filename

        db.execute(
            """INSERT INTO payments (project_id, contract_id, participant_id, payment_no, amount,
               payment_date, payment_method, payer, remark, attachment)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (pid, contract_id, participant_id, payment_no, amount, payment_date,
             payment_method, payer, remark, attachment)
        )
        db.commit()
        add_log(session['user_id'], session['username'], '新增付款',
                f'项目{pid}新增付款: {payment_no} {amount}元', request.remote_addr)
        flash('付款记录添加成功', 'success')
        return redirect(url_for('project_detail', pid=pid))

    return render_template('payment_form.html', pid=pid, contracts=contracts, participants=participants, now=datetime.now())


# ==================== 路由：往来款管理（更新版） ====================

@app.route('/project/<int:pid>/fund/add', methods=['GET', 'POST'])
@login_required
def fund_transaction_add(pid):
    db = get_db()
    participants = db.execute("""
        SELECT p.* FROM participants p
        JOIN project_participants pp ON p.id = pp.participant_id
        WHERE pp.project_id=?
    """, (pid,)).fetchall()

    if request.method == 'POST':
        participant_id = request.form.get('participant_id') or None
        trans_type = request.form.get('trans_type', '入账')
        amount = float(request.form.get('amount', 0) or 0)
        trans_date = request.form.get('trans_date', '')
        payment_method = request.form.get('payment_method', '')
        payer = request.form.get('payer', '')
        remark = request.form.get('remark', '')

        # 处理附件上传
        attachment = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename:
                filename = f"fund_{uuid.uuid4().hex[:8]}_{file.filename}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                attachment = filename

        db.execute(
            """INSERT INTO transactions (project_id, participant_id, trans_type, amount, trans_date,
               payment_method, payer, remark, attachment)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (pid, participant_id, trans_type, amount, trans_date, payment_method, payer, remark, attachment)
        )
        db.commit()
        add_log(session['user_id'], session['username'], '新增往来款',
                f'项目{pid}新增{trans_type}: {amount}元', request.remote_addr)
        flash('往来款记录添加成功', 'success')
        return redirect(url_for('project_detail', pid=pid))

    return render_template('fund_form.html', pid=pid, participants=participants, now=datetime.now())


# ==================== 路由：图片 OCR 识别 ====================

@app.route('/ocr/upload', methods=['GET', 'POST'])
@login_required
def ocr_upload():
    if request.method == 'POST':
        if 'image' not in request.files:
            flash('请选择图片文件', 'warning')
            return redirect(url_for('ocr_upload'))

        file = request.files['image']
        if not file.filename:
            flash('请选择图片文件', 'warning')
            return redirect(url_for('ocr_upload'))

        # 检查文件类型
        allowed_ext = ('.jpg', '.jpeg', '.png', '.webp')
        if not file.filename.lower().endswith(allowed_ext):
            flash('仅支持 JPG、PNG、WEBP 格式图片', 'warning')
            return redirect(url_for('ocr_upload'))

        # 保存图片
        filename = f"ocr_{uuid.uuid4().hex[:12]}_{file.filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr', filename)
        file.save(filepath)

        # OCR 识别
        try:
            transactions, raw_text = ocr_recognize(filepath)

            # 保存到草稿表
            draft_data = {
                'image_path': filename,
                'transactions': transactions,
                'raw_text': raw_text
            }
            db = get_db()
            cursor = db.execute(
                "INSERT INTO ocr_drafts (image_path, draft_data) VALUES (?, ?)",
                (filename, json.dumps(draft_data, ensure_ascii=False))
            )
            draft_id = cursor.lastrowid
            db.commit()

            return redirect(url_for('ocr_review', draft_id=draft_id))

        except Exception as e:
            flash(f'识别失败: {str(e)}', 'danger')
            return redirect(url_for('ocr_upload'))

    return render_template('ocr_upload.html')


@app.route('/ocr/review/<int:draft_id>')
@login_required
def ocr_review(draft_id):
    db = get_db()
    draft = db.execute("SELECT * FROM ocr_drafts WHERE id=?", (draft_id,)).fetchone()

    if not draft:
        flash('草稿不存在', 'danger')
        return redirect(url_for('ocr_upload'))

    draft_data = json.loads(draft['draft_data'])
    projects = db.execute("SELECT * FROM projects ORDER BY name").fetchall()
    participants = db.execute("SELECT * FROM participants ORDER BY name").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY type, name").fetchall()

    return render_template('ocr_review.html',
                           draft=draft_data,
                           draft_id=draft_id,
                           projects=projects,
                           participants=participants,
                           categories=categories)


@app.route('/ocr/confirm/<int:draft_id>', methods=['POST'])
@login_required
def ocr_confirm(draft_id):
    """确认 OCR 识别结果并保存"""
    db = get_db()
    draft = db.execute("SELECT * FROM ocr_drafts WHERE id=?", (draft_id,)).fetchone()

    if not draft:
        return jsonify({'success': False, 'message': '草稿不存在'})

    try:
        data = request.get_json()
        records = data.get('records', [])

        saved_count = 0
        for record in records:
            if not record.get('selected'):
                continue

            project_id = record.get('project_id')
            if not project_id:
                continue

            ocr_att = draft['image_path']
            if ocr_att and not ocr_att.replace('\\', '/').startswith('ocr/'):
                ocr_att = f"ocr/{ocr_att}"
            db.execute(
                """INSERT INTO transaction_records
                   (project_id, participant_id, category_id, trans_date, amount, trans_type,
                    description, merchant, payment_method, source, attachment)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'image-ocr', ?)""",
                (project_id,
                 record.get('participant_id'),
                 record.get('category_id'),
                 record.get('date', datetime.now().strftime('%Y-%m-%d')),
                 float(record.get('amount', 0)),
                 record.get('trans_type', 'expense'),
                 record.get('description', ''),
                 record.get('merchant', ''),
                 record.get('payment_method', ''),
                 ocr_att)
            )
            saved_count += 1

        db.commit()

        # 更新草稿状态
        db.execute("UPDATE ocr_drafts SET status='confirmed' WHERE id=?", (draft_id,))
        db.commit()

        add_log(session['user_id'], session['username'], 'OCR确认入账',
                f'确认{saved_count}条记录', request.remote_addr)

        return jsonify({'success': True, 'saved': saved_count})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ==================== 路由：Excel 导入 ====================

@app.route('/import/excel', methods=['GET', 'POST'])
@login_required
def excel_import():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择 Excel 文件', 'warning')
            return redirect(url_for('excel_import'))

        file = request.files['file']
        if not file.filename:
            flash('请选择 Excel 文件', 'warning')
            return redirect(url_for('excel_import'))

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('仅支持 Excel 文件 (.xlsx, .xls)', 'warning')
            return redirect(url_for('excel_import'))

        # 保存文件
        filename = f"import_{uuid.uuid4().hex[:8]}_{file.filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # 解析 Excel
        records = parse_excel(filepath)

        if isinstance(records, dict) and 'error' in records:
            flash(f'解析失败: {records["error"]}', 'danger')
            return redirect(url_for('excel_import'))

        # 导入数据
        project_name = request.form.get('project_name', '').strip()
        results = import_from_excel(records, project_name, session['user_id'])

        add_log(session['user_id'], session['username'], 'Excel导入',
                f'导入{results["imported"]}条, 更新{results["updated"]}条', request.remote_addr)

        flash(f"""导入完成！
新建项目: {results['project_created']}个
新建参与人: {results['participant_created']}人
导入记录: {results['imported']}条
更新记录: {results['updated']}条
跳过: {results['skipped']}条""", 'success')

        return redirect(url_for('transaction_records'))

    return render_template('excel_import.html')


# ==================== 路由：费用分类（标准库维护） ====================

@app.route('/categories')
@login_required
def category_list():
    from route_extensions import build_category_manage_trees
    from project_category_utils import fetch_leaf_categories
    db = get_db()
    expense_tree, income_tree, all_categories, usage_map = build_category_manage_trees(db)
    total_count = db.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
    leaf_count = len(fetch_leaf_categories(db))
    total_usage = sum(usage_map.values()) if usage_map else 0
    return render_template('category_list.html',
                           expense_tree=expense_tree,
                           income_tree=income_tree,
                           all_categories=all_categories,
                           total_count=total_count,
                           leaf_count=leaf_count,
                           total_usage=total_usage)


@app.route('/category/add', methods=['POST'])
@login_required
def category_add():
    name = request.form.get('name', '').strip()
    code = request.form.get('code', '').strip()
    type_ = request.form.get('type', 'expense')
    level = request.form.get('level', type=int) or 1
    parent_id = request.form.get('parent_id', type=int) or None
    sort_order = request.form.get('sort_order', type=int) or 0
    description = request.form.get('description', '').strip()

    if not name or not code:
        flash('科目名称和科目代码不能为空', 'warning')
        return redirect(url_for('category_list'))

    db = get_db()
    dup = db.execute("SELECT 1 FROM categories WHERE code=?", (code,)).fetchone()
    if dup:
        flash('科目代码已存在', 'warning')
        return redirect(url_for('category_list'))

    if level > 1 and not parent_id:
        flash('二级、三级科目必须选择上级科目', 'warning')
        return redirect(url_for('category_list'))
    if level == 1:
        parent_id = None
    if parent_id:
        parent = db.execute("SELECT level, type FROM categories WHERE id=?", (parent_id,)).fetchone()
        if not parent:
            flash('上级科目不存在', 'warning')
            return redirect(url_for('category_list'))
        if parent['type'] != type_:
            flash('上级科目与当前类型（收入/支出）不一致', 'warning')
            return redirect(url_for('category_list'))
        if (parent['level'] or 1) != level - 1:
            flash('上级科目层级不匹配', 'warning')
            return redirect(url_for('category_list'))

    db.execute(
        """INSERT INTO categories (name, code, type, level, parent_id, sort_order, description, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (name, code, type_, level, parent_id, sort_order, description,
         datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    )
    db.commit()
    flash('费用分类添加成功', 'success')
    return redirect(url_for('category_list'))


@app.route('/category/<int:cid>/edit', methods=['GET', 'POST'])
@login_required
def category_edit(cid):
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE id=?", (cid,)).fetchone()
    if not cat:
        flash('科目不存在', 'danger')
        return redirect(url_for('category_list'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip()
        sort_order = request.form.get('sort_order', type=int) or 0
        description = request.form.get('description', '').strip()
        if not name or not code:
            flash('科目名称和代码不能为空', 'warning')
            return redirect(url_for('category_edit', cid=cid))
        dup = db.execute(
            "SELECT 1 FROM categories WHERE code=? AND id!=?", (code, cid)
        ).fetchone()
        if dup:
            flash('科目代码已被其他科目使用', 'warning')
            return redirect(url_for('category_edit', cid=cid))
        db.execute(
            """UPDATE categories SET name=?, code=?, sort_order=?, description=?
               WHERE id=?""",
            (name, code, sort_order, description, cid),
        )
        db.commit()
        flash('费用分类已更新', 'success')
        return redirect(url_for('category_list'))

    from route_extensions import build_category_manage_trees
    _, _, all_categories, _ = build_category_manage_trees(db)
    usage = db.execute(
        "SELECT COUNT(*) FROM transaction_records WHERE category_id=?", (cid,)
    ).fetchone()[0]
    child_count = db.execute(
        "SELECT COUNT(*) FROM categories WHERE parent_id=?", (cid,)
    ).fetchone()[0]
    return render_template('category_form.html',
                           category=cat,
                           all_categories=all_categories,
                           usage_count=usage,
                           child_count=child_count,
                           edit_mode=True)


# ==================== 路由：账号管理 ====================

@app.route('/account')
@login_required
@admin_required
def account_manage():
    db = get_db()
    keyword = request.args.get('keyword', '').strip()
    role_filter = request.args.get('role', '').strip()
    status_filter = request.args.get('status', '').strip()
    try:
        sql = (
            "SELECT id, username, real_name, phone, email, department, role, "
            "COALESCE(status, 'active') as status, created_at, last_login "
            "FROM users WHERE 1=1"
        )
        params = []
        if keyword:
            sql += (
                " AND (username LIKE ? OR real_name LIKE ? OR phone LIKE ? "
                "OR email LIKE ? OR department LIKE ?)"
            )
            like = f'%{keyword}%'
            params.extend([like, like, like, like, like])
        if role_filter:
            sql += " AND role=?"
            params.append(role_filter)
        if status_filter:
            sql += " AND COALESCE(status, 'active')=?"
            params.append(status_filter)
        sql += " ORDER BY id"
        users = db.execute(sql, params).fetchall()
        all_users = db.execute(
            "SELECT id, username, real_name, phone, email, department, role, "
            "COALESCE(status, 'active') as status, created_at, last_login "
            "FROM users ORDER BY id"
        ).fetchall()
    except sqlite3.OperationalError:
        users = db.execute(
            "SELECT id, username, role, created_at FROM users ORDER BY id"
        ).fetchall()
        all_users = users
    def _user_status(row):
        return row['status'] if hasattr(row, 'keys') and 'status' in row.keys() and row['status'] else 'active'
    stats = {
        'total': len(all_users),
        'active': sum(1 for u in all_users if _user_status(u) == 'active'),
        'pending': sum(1 for u in all_users if _user_status(u) == 'pending'),
        'disabled': sum(1 for u in all_users if _user_status(u) == 'disabled'),
    }
    client_accounts = db.execute(
        """SELECT ca.*, c.name as customer_master_name, u.username as approver_name
           FROM client_accounts ca
           LEFT JOIN customers c ON ca.customer_id = c.id
           LEFT JOIN users u ON ca.approved_by = u.id
           ORDER BY ca.created_at DESC"""
    ).fetchall()
    customers = db.execute(
        "SELECT id, name FROM customers WHERE is_active=1 ORDER BY name"
    ).fetchall()
    return render_template(
        'account_manage.html',
        users=users,
        roles=STAFF_ROLES,
        client_accounts=client_accounts,
        customers=customers,
        filters={'keyword': keyword, 'role': role_filter, 'status': status_filter},
        stats=stats,
    )


@app.route('/admin/users/<int:uid>/collab-scope', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_user_collab_scope(uid):
    """管理员：为协同专员分配可管理的客户公司。"""
    db = get_db()
    user = db.execute('SELECT id, username, role FROM users WHERE id=?', (uid,)).fetchone()
    if not user:
        flash('用户不存在', 'danger')
        return redirect(url_for('account_manage'))
    if user['role'] != ROLE_CLIENT_COLLAB:
        flash('仅可为「客户协同专员」分配负责客户', 'warning')
        return redirect(url_for('account_manage'))
    if request.method == 'POST':
        ids = request.form.getlist('customer_ids')
        set_user_assignments(db, uid, ids, session.get('user_id'))
        db.commit()
        flash(f'已为 {user["username"]} 更新负责客户（{len(ids)} 家）', 'success')
        return redirect(url_for('account_manage'))
    customers = db.execute(
        'SELECT id, name FROM customers WHERE is_active=1 ORDER BY name'
    ).fetchall()
    assigned = {r['customer_id'] for r in get_user_assignments(db, uid)}
    return render_template(
        'admin_user_collab_scope.html',
        user=user,
        customers=customers,
        assigned=assigned,
    )


@app.route('/account/add', methods=['POST'])
@login_required
@admin_required
def account_add():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    role = request.form.get('role', 'user')
    real_name = request.form.get('real_name', '').strip()
    phone = request.form.get('phone', '').strip()
    email = request.form.get('email', '').strip()
    department = request.form.get('department', '').strip()
    allowed_roles = {r['code'] for r in STAFF_ROLES}
    if role not in allowed_roles:
        role = 'user'

    if not username or not password:
        flash('用户名和密码不能为空', 'warning')
        return redirect(url_for('account_manage'))

    hashed = hashlib.md5(password.encode()).hexdigest()
    db = get_db()
    try:
        try:
            db.execute(
                """INSERT INTO users
                   (username, password, real_name, phone, email, department, role, status, created_by)
                   VALUES (?, ?, ?, ?, ?, ?, ?, 'active', ?)""",
                (username, hashed, real_name, phone, email, department, role, session.get('user_id')),
            )
        except sqlite3.OperationalError:
            db.execute(
                "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                (username, hashed, role),
            )
        db.commit()
        add_log(session['user_id'], session['username'], '新增账号', f'新增用户: {username}({role})', request.remote_addr)
        if role == ROLE_CLIENT_COLLAB:
            flash(f'客户协同专员「{username}」已创建，登录后仅可管理客户协同模块', 'success')
        else:
            flash('账号创建成功', 'success')
    except sqlite3.IntegrityError:
        flash('用户名已存在', 'danger')
    return redirect(url_for('account_manage'))


@app.route('/account/<int:uid>/update', methods=['POST'])
@login_required
@admin_required
def account_update(uid):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user:
        flash('账号不存在', 'danger')
        return redirect(url_for('account_manage'))

    username = request.form.get('username', '').strip()
    real_name = request.form.get('real_name', '').strip()
    phone = request.form.get('phone', '').strip()
    email = request.form.get('email', '').strip()
    department = request.form.get('department', '').strip()
    role = request.form.get('role', user['role'] or 'user')
    status = request.form.get('status', user['status'] or 'active')
    allowed_roles = {r['code'] for r in STAFF_ROLES}
    allowed_status = {'active', 'pending', 'disabled'}
    if not username:
        flash('用户名不能为空', 'warning')
        return redirect(url_for('account_manage'))
    if role not in allowed_roles:
        role = user['role'] or 'user'
    if status not in allowed_status:
        status = user['status'] or 'active'
    if uid == session.get('user_id') and status != 'active':
        flash('不能禁用或挂起当前登录账号', 'danger')
        return redirect(url_for('account_manage'))
    if user['role'] == 'admin' and role != 'admin':
        admin_count = db.execute(
            "SELECT COUNT(*) FROM users WHERE role='admin' AND COALESCE(status, 'active')='active'"
        ).fetchone()[0]
        if admin_count <= 1:
            flash('至少保留一个可用管理员账号', 'danger')
            return redirect(url_for('account_manage'))

    try:
        db.execute(
            """UPDATE users
               SET username=?, real_name=?, phone=?, email=?, department=?, role=?, status=?
               WHERE id=?""",
            (username, real_name, phone, email, department, role, status, uid),
        )
        db.commit()
        add_log(session['user_id'], session['username'], '维护账号',
                f'更新用户: {username}({role}/{status})', request.remote_addr)
        flash('账号资料已更新', 'success')
    except sqlite3.IntegrityError:
        flash('用户名已存在', 'danger')
    return redirect(url_for('account_manage'))


@app.route('/account/<int:uid>/reset-password', methods=['POST'])
@login_required
@admin_required
def account_reset_password(uid):
    password = request.form.get('password', '').strip()
    if len(password) < 4:
        flash('密码至少 4 位', 'warning')
        return redirect(url_for('account_manage'))
    db = get_db()
    user = db.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    if not user:
        flash('账号不存在', 'danger')
        return redirect(url_for('account_manage'))
    hashed = hashlib.md5(password.encode()).hexdigest()
    db.execute("UPDATE users SET password=? WHERE id=?", (hashed, uid))
    db.commit()
    add_log(session['user_id'], session['username'], '重置账号密码',
            f'用户: {user["username"]}', request.remote_addr)
    flash(f'{user["username"]} 的密码已重置', 'success')
    return redirect(url_for('account_manage'))


@app.route('/account/delete/<int:uid>')
@login_required
@admin_required
def account_delete(uid):
    if uid == session['user_id']:
        flash('不能删除当前登录账号', 'danger')
        return redirect(url_for('account_manage'))

    db = get_db()
    user = db.execute("SELECT username, role FROM users WHERE id=?", (uid,)).fetchone()
    if user and user['role'] == 'admin':
        admin_count = db.execute(
            "SELECT COUNT(*) FROM users WHERE role='admin' AND COALESCE(status, 'active')='active'"
        ).fetchone()[0]
        if admin_count <= 1:
            flash('至少保留一个可用管理员账号', 'danger')
            return redirect(url_for('account_manage'))
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.commit()
    add_log(session['user_id'], session['username'], '删除账号', f'删除用户ID: {uid}', request.remote_addr)
    flash('账号已删除', 'success')
    return redirect(url_for('account_manage'))


# ==================== 路由：数据导入/导出 ====================

@app.route('/account/export')
@login_required
@admin_required
def data_export():
    db = get_db()
    data = {}

    tables = ['users', 'projects', 'participants', 'categories', 'contracts', 'invoices',
              'payments', 'transactions', 'transaction_records', 'investments', 'dividends', 'logs']
    for table in tables:
        try:
            rows = db.execute(f"SELECT * FROM {table}").fetchall()
            data[table] = [dict(row) for row in rows]
        except:
            pass

    filepath = os.path.join(app.config['BACKUP_DIR'], f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json')
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    add_log(session['user_id'], session['username'], '导出数据', '导出全部数据为JSON', request.remote_addr)
    return send_file(filepath, as_attachment=True, download_name='project_data_export.json')


@app.route('/account/clear')
@login_required
@admin_required
def data_clear():
    db = get_db()
    tables = ['logs', 'transaction_records', 'transactions', 'payments', 'invoices',
              'contracts', 'investments', 'dividends', 'project_participants', 'projects']
    for table in tables:
        try:
            db.execute(f"DELETE FROM {table}")
        except:
            pass
    db.commit()
    add_log(session['user_id'], session['username'], '清空数据', '清空所有业务数据', request.remote_addr)
    flash('数据已清空（保留用户账号和参与人）', 'success')
    return redirect(url_for('account_manage'))


# ==================== 路由：备份管理 ====================

@app.route('/backup')
@login_required
@admin_required
def backup_manage():
    backups = []
    backup_dir = app.config['BACKUP_DIR']
    if os.path.exists(backup_dir):
        for f in sorted(os.listdir(backup_dir), reverse=True):
            if f.endswith('.db') or f.endswith('.json'):
                fpath = os.path.join(backup_dir, f)
                stat = os.stat(fpath)
                backups.append({
                    'name': f,
                    'size': round(stat.st_size / 1024, 1),
                    'time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
    return render_template('backup.html', backups=backups)


@app.route('/backup/manual')
@login_required
@admin_required
def backup_manual():
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = os.path.join(app.config['BACKUP_DIR'], f'backup_{timestamp}.db')
    shutil.copy2(app.config['DATABASE'], backup_file)
    add_log(session['user_id'], session['username'], '手动备份', f'备份数据库: backup_{timestamp}.db', request.remote_addr)
    flash('手动备份成功', 'success')
    return redirect(url_for('backup_manage'))


@app.route('/backup/download/<filename>')
@login_required
@admin_required
def backup_download(filename):
    filepath = os.path.join(app.config['BACKUP_DIR'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    flash('文件不存在', 'danger')
    return redirect(url_for('backup_manage'))


@app.route('/backup/delete/<filename>')
@login_required
@admin_required
def backup_delete(filename):
    filepath = os.path.join(app.config['BACKUP_DIR'], filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        add_log(session['user_id'], session['username'], '删除备份', f'删除备份文件: {filename}', request.remote_addr)
        flash('备份文件已删除', 'success')
    return redirect(url_for('backup_manage'))


# ==================== 路由：操作日志 ====================

@app.route('/logs')
@login_required
def log_list():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page

    total = db.execute("SELECT COUNT(*) FROM logs").fetchone()[0]
    logs = db.execute(
        "SELECT * FROM logs ORDER BY created_at DESC LIMIT ? OFFSET ?",
        (per_page, offset)
    ).fetchall()

    total_pages = (total + per_page - 1) // per_page

    return render_template('log_list.html', logs=logs, page=page, total_pages=total_pages, total=total)


# ==================== 定时备份 ====================

def run_scheduler():
    """后台定时备份任务"""
    schedule.every().day.at("02:00").do(do_scheduled_backup)
    while True:
        schedule.run_pending()
        time.sleep(3600)


def do_scheduled_backup():
    """执行定时备份"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(app.config['BACKUP_DIR'], f'auto_backup_{timestamp}.db')
        shutil.copy2(app.config['DATABASE'], backup_file)
    except Exception:
        pass


# ==================== 启动 ====================

# ==================== 客户协同模块 ====================

def client_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'client_id' not in session:
            return redirect(url_for('portal_login'))
        return f(*args, **kwargs)
    return decorated_function


def check_balance_alert(db, client):
    """检查余额预警"""
    if not client or client['total_recharge'] <= 0:
        return
    threshold = client['alert_threshold'] or 10
    if client['balance'] < client['total_recharge'] * threshold / 100:
        now = datetime.now()
        if client['last_alert_at']:
            last = datetime.strptime(client['last_alert_at'], '%Y-%m-%d %H:%M:%S')
            if (now - last).total_seconds() < 3600:
                return
        db.execute("""INSERT INTO client_messages (client_id, title, content, msg_type)
                     VALUES (?, ?, ?, 'alert')""",
                   (client['id'], '余额不足提醒',
                    f'您的账户余额为 {client["balance"]:.2f} 元，已低于充值总额的 {threshold}%，请及时充值！'))
        db.execute("UPDATE client_accounts SET last_alert_at=? WHERE id=?",
                   (now.strftime('%Y-%m-%d %H:%M:%S'), client['id']))
        db.commit()


# ---------- 客户端路由 ----------

@app.route('/portal/login', methods=['GET', 'POST'])
def portal_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if not username or not password:
            flash('请输入用户名和密码', 'danger')
            return render_template('portal_login.html')
        db = get_db()
        client = db.execute("SELECT * FROM client_accounts WHERE username=?", (username,)).fetchone()
        if not client or not check_password_hash(client['password'], password):
            flash('用户名或密码错误', 'danger')
            return render_template('portal_login.html')
        st = normalize_client_status(client['status'])
        if st == CLIENT_STATUS_PENDING:
            flash('您的账户正在审核中，请耐心等待', 'warning')
            return render_template('portal_login.html')
        if st == CLIENT_STATUS_REJECTED:
            flash('您的账户审核未通过，请联系管理员', 'danger')
            return render_template('portal_login.html')
        if st == CLIENT_STATUS_DISABLED:
            flash('账户已禁用，请联系管理员', 'danger')
            return render_template('portal_login.html')
        if st != CLIENT_STATUS_APPROVED:
            flash('账户状态异常，请联系管理员', 'danger')
            return render_template('portal_login.html')
        session['client_id'] = client['id']
        session['client_name'] = client['contact_name'] or client['username']
        session['client_company_name'] = client['company_name'] or ''
        session['username'] = client['username']
        session['customer_id'] = client['customer_id']
        bal, _, _ = aggregate_company_balance(db, dict(client))
        session['client_balance'] = bal
        scope_ids = company_scope_client_ids(db, dict(client))
        if scope_ids:
            ph = ','.join('?' * len(scope_ids))
            session['unread_count'] = db.execute(
                f"SELECT COUNT(*) FROM client_messages WHERE client_id IN ({ph}) AND is_read=0",
                scope_ids,
            ).fetchone()[0]
        else:
            session['unread_count'] = 0
        add_log(None, username, '客户登录', f"客户 {username} 登录系统")
        return redirect(url_for('portal_index'))
    return render_template('portal_login.html')


@app.route('/portal/register', methods=['GET', 'POST'])
def portal_register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        phone = request.form.get('phone', '').strip()
        # 模板字段名为 contact_person，兼容 contact_name
        contact_name = (
            request.form.get('contact_name', '').strip()
            or request.form.get('contact_person', '').strip()
        )
        company_name = request.form.get('company_name', '').strip()
        if not all([username, password, phone, contact_name, company_name]):
            flash('请填写所有必填项', 'danger')
            return render_template('portal_register.html')
        if password != confirm:
            flash('两次密码输入不一致', 'danger')
            return render_template('portal_register.html')
        if len(password) < 6:
            flash('密码长度不能少于6位', 'danger')
            return render_template('portal_register.html')
        db = get_db()
        existing = db.execute("SELECT id FROM client_accounts WHERE username=?", (username,)).fetchone()
        if existing:
            flash('用户名已存在', 'danger')
            return render_template('portal_register.html')
        customer_id = resolve_or_create_customer(
            db, company_name, contact_name, phone)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        db.execute("""INSERT INTO client_accounts
                     (customer_id, username, password, phone, contact_name, company_name,
                      status, balance, total_recharge, total_deduct, created_at, updated_at)
                     VALUES (?, ?, ?, ?, ?, ?, 'pending', 0, 0, 0, ?, ?)""",
                   (customer_id, username, generate_password_hash(password),
                    phone, contact_name, company_name, now, now))
        db.commit()
        add_log(None, username, '客户注册', f"新客户 {username}({company_name}) 注册申请")
        flash('注册成功，请等待管理员审核', 'success')
        return redirect(url_for('portal_login'))
    return render_template('portal_register.html')


@app.route('/portal/logout')
def portal_logout():
    for key in ('client_id', 'client_name', 'client_company_name', 'client_balance', 'customer_id'):
        session.pop(key, None)
    flash('已退出登录', 'info')
    return redirect(url_for('portal_login'))


def _refresh_portal_session(db, client):
    """同步门户 session 展示字段"""
    client_d = dict(client)
    bal, total_recharge, total_deduct = aggregate_company_balance(db, client_d)
    session['client_balance'] = bal
    session['client_company_name'] = client_d.get('company_name') or ''
    session['username'] = client_d.get('username') or session.get('username', '')
    scope_ids = company_scope_client_ids(db, client_d)
    if scope_ids:
        ph = ','.join('?' * len(scope_ids))
        session['unread_count'] = db.execute(
            f"SELECT COUNT(*) FROM client_messages WHERE client_id IN ({ph}) AND is_read=0",
            scope_ids,
        ).fetchone()[0]
    else:
        session['unread_count'] = 0
    return bal, total_recharge, total_deduct, scope_ids


@app.route('/portal/')
@client_login_required
def portal_index():
    db = get_db()
    client = db.execute("SELECT * FROM client_accounts WHERE id=?", (session['client_id'],)).fetchone()
    if not client:
        flash('账户不存在', 'danger')
        return redirect(url_for('portal_logout'))
    check_balance_alert(db, client)
    client = db.execute("SELECT * FROM client_accounts WHERE id=?", (session['client_id'],)).fetchone()
    balance, total_recharge, total_deduction, scope_ids = _refresh_portal_session(db, client)

    unread_messages = []
    recent_transactions = []
    if scope_ids:
        ph = ','.join('?' * len(scope_ids))
        unread_messages = db.execute(
            f"""SELECT * FROM client_messages WHERE client_id IN ({ph}) AND is_read=0
                ORDER BY created_at DESC LIMIT 10""",
            scope_ids,
        ).fetchall()
    recent_transactions, _, _, _ = build_portal_transactions(db, dict(client))
    recent_transactions = recent_transactions[:5]

    filt, fparams = sales_order_customer_filter(dict(client))
    recent_deliveries = db.execute(
        f"""SELECT so.id, so.order_no as order_number, so.order_date, so.status,
                   COALESCE(so.total_amount, 0) as total_amount,
                   (SELECT COUNT(*) FROM sales_order_items si WHERE si.sales_order_id = so.id) as item_count
            FROM sales_orders so WHERE {filt}
            ORDER BY so.created_at DESC LIMIT 5""",
        fparams,
    ).fetchall()

    return render_template(
        'portal_index.html',
        client=client,
        balance=balance,
        total_recharge=total_recharge,
        total_deduction=total_deduction,
        recent_transactions=recent_transactions,
        unread_messages=unread_messages,
        recent_deliveries=recent_deliveries,
    )


@app.route('/portal/recharge', methods=['GET', 'POST'])
@client_login_required
def portal_recharge():
    if request.method == 'POST':
        amount = request.form.get('amount', '').strip()
        payment_method = request.form.get('payment_method', '').strip()
        payment_no = request.form.get('payment_no', '').strip()
        remark = request.form.get('remark', '').strip()
        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError
        except (ValueError, TypeError):
            flash('请输入有效的充值金额', 'danger')
            return render_template('portal_recharge.html')
        if not payment_method:
            flash('请选择付款方式', 'danger')
            return render_template('portal_recharge.html')
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        db.execute("""INSERT INTO client_recharges (client_id, amount, payment_method, payment_no, remark, status, created_at)
                     VALUES (?, ?, ?, ?, ?, 'pending', ?)""",
                   (session['client_id'], amount, payment_method, payment_no, remark, now))
        db.commit()
        add_log(None, session.get('client_name', ''), '申请充值',
                f"客户申请充值 {amount:.2f} 元")
        flash('充值申请已提交，请等待管理员确认', 'success')
        return redirect(url_for('portal_recharges'))
    return render_template('portal_recharge.html')


@app.route('/portal/recharges')
@client_login_required
def portal_recharges():
    db = get_db()
    client = get_client_by_id(db, session['client_id'])
    if client:
        _refresh_portal_session(db, client)
    scope_ids = company_scope_client_ids(db, client)
    ph = ','.join('?' * len(scope_ids))
    recharges = db.execute(
        f"SELECT * FROM client_recharges WHERE client_id IN ({ph}) ORDER BY created_at DESC",
        scope_ids,
    ).fetchall()
    return render_template('portal_recharges.html', recharges=recharges)


@app.route('/portal/deliveries')
@client_login_required
def portal_deliveries():
    db = get_db()
    client = get_client_by_id(db, session['client_id'])
    if client:
        _refresh_portal_session(db, client)
    filt, fparams = sales_order_customer_filter(client)
    deliveries = db.execute(
        f"""SELECT so.id, so.order_no as order_number, so.order_date, so.status,
                   COALESCE(so.total_amount, 0) as total_amount,
                   (SELECT COUNT(*) FROM sales_order_items si WHERE si.sales_order_id = so.id) as item_count,
                   c.name as customer_name
            FROM sales_orders so
            LEFT JOIN customers c ON so.customer_id = c.id
            WHERE {filt}
            ORDER BY so.created_at DESC""",
        fparams,
    ).fetchall()
    return render_template('portal_deliveries.html', deliveries=deliveries)


@app.route('/portal/delivery/<int:id>')
@client_login_required
def portal_delivery_detail(id):
    db = get_db()
    client = get_client_by_id(db, session['client_id'])
    filt, fparams = sales_order_customer_filter(client)
    order = db.execute(
        f"""SELECT so.*, c.name as customer_name
            FROM sales_orders so
            LEFT JOIN customers c ON so.customer_id = c.id
            WHERE so.id=? AND {filt}""",
        (id,) + fparams,
    ).fetchone()
    if not order:
        flash('订单不存在或无权查看', 'danger')
        return redirect(url_for('portal_deliveries'))
    items = db.execute(
        """SELECT si.*, si.item_name as product_name, si.specification
           FROM sales_order_items si WHERE si.sales_order_id=?""",
        (id,),
    ).fetchall()
    return render_template('portal_delivery_detail.html', delivery=order, items=items)


@app.route('/portal/transactions')
@client_login_required
def portal_transactions():
    db = get_db()
    client = get_client_by_id(db, session['client_id'])
    tx_type = request.args.get('type', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    transactions, current_balance, total_recharge, total_deduction = build_portal_transactions(
        db, client, tx_type, start_date, end_date)
    return render_template(
        'portal_transactions.html',
        transactions=transactions,
        current_balance=current_balance,
        total_recharge=total_recharge,
        total_deduction=total_deduction,
    )


@app.route('/portal/messages')
@client_login_required
def portal_messages():
    db = get_db()
    client = get_client_by_id(db, session['client_id'])
    scope_ids = company_scope_client_ids(db, client)
    ph = ','.join('?' * len(scope_ids))
    messages = db.execute(
        f"SELECT * FROM client_messages WHERE client_id IN ({ph}) ORDER BY created_at DESC",
        scope_ids,
    ).fetchall()
    return render_template('portal_messages.html', messages=messages)


@app.route('/portal/message/<int:id>/read')
@client_login_required
def portal_message_read(id):
    db = get_db()
    client = get_client_by_id(db, session['client_id'])
    scope_ids = company_scope_client_ids(db, client)
    ph = ','.join('?' * len(scope_ids))
    msg = db.execute(
        f"SELECT * FROM client_messages WHERE id=? AND client_id IN ({ph})",
        (id,) + tuple(scope_ids),
    ).fetchone()
    if msg:
        if not msg['is_read']:
            db.execute("UPDATE client_messages SET is_read=1 WHERE id=?", (id,))
            db.commit()
        return render_template('portal_message_detail.html', message=msg)
    flash('消息不存在', 'danger')
    return redirect(url_for('portal_messages'))


@app.route('/portal/messages/read-all', methods=['POST'])
@client_login_required
def portal_messages_read_all():
    db = get_db()
    db.execute("UPDATE client_messages SET is_read=1 WHERE client_id=? AND is_read=0",
               (session['client_id'],))
    db.commit()
    flash('已全部标记为已读', 'success')
    return redirect(url_for('portal_messages'))


@app.route('/portal/report')
@client_login_required
def portal_report():
    db = get_db()
    client = get_client_by_id(db, session['client_id'])
    balance, total_recharge, total_deduct = aggregate_company_balance(db, client)
    scope_ids = company_scope_client_ids(db, client)
    ph = ','.join('?' * len(scope_ids))
    deductions = db.execute(
        f"""SELECT item_name, SUM(quantity) as total_qty,
                   SUM(amount) as total_amount, unit_price
            FROM client_deductions WHERE client_id IN ({ph})
            GROUP BY item_name ORDER BY total_amount DESC""",
        scope_ids,
    ).fetchall()
    monthly = db.execute(
        f"""SELECT strftime('%Y-%m', COALESCE(deduct_date, created_at)) as month,
                   SUM(amount) as month_amount
            FROM client_deductions WHERE client_id IN ({ph})
            GROUP BY month ORDER BY month DESC LIMIT 12""",
        scope_ids,
    ).fetchall()
    client_view = dict(client)
    client_view['balance'] = balance
    client_view['total_recharge'] = total_recharge
    client_view['total_deduct'] = total_deduct
    return render_template(
        'portal_report.html', client=client_view,
        total_recharge=total_recharge, total_deduct=total_deduct, balance=balance,
        deductions=deductions, monthly=monthly,
    )


# ---------- 管理端客户协同路由 ----------

@app.route('/admin/client-dashboard')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_dashboard():
    db = get_db()
    q = request.args.get('q', '').strip()
    uid, role = session.get('user_id'), session.get('role')
    rows = list_company_summaries_scoped(db, uid, role)
    if q:
        rows = [r for r in rows if q.lower() in (r.get('company_name') or '').lower()]
    total_balance = sum(float(r.get('balance') or 0) for r in rows)
    total_recharge = sum(float(r.get('total_recharge') or 0) for r in rows)
    total_deduct = sum(float(r.get('total_deduct') or 0) for r in rows)
    return render_template(
        'admin_client_dashboard.html',
        companies=rows,
        total_balance=total_balance,
        total_recharge=total_recharge,
        total_deduct=total_deduct,
        q=q,
    )


@app.route('/admin/client-company/<int:customer_id>')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_workspace(customer_id):
    db = get_db()
    denied = assert_customer_access(
        db, session.get('user_id'), session.get('role'), customer_id)
    if denied:
        return denied
    workspace = get_company_workspace(db, customer_id=customer_id)
    if not workspace:
        flash('未找到该客户公司或未绑定激活账号', 'warning')
        return redirect(url_for('admin_client_dashboard'))
    return render_template('admin_client_workspace.html', ws=workspace)


@app.route('/admin/client-company/<int:customer_id>/recharge', methods=['POST'])
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_company_recharge(customer_id):
    db = get_db()
    denied = assert_customer_access(
        db, session.get('user_id'), session.get('role'), customer_id)
    if denied:
        return denied
    workspace = get_company_workspace(db, customer_id=customer_id)
    if not workspace:
        flash('公司不存在或未绑定激活账号', 'danger')
        return redirect(url_for('admin_client_dashboard'))
    amount = request.form.get('amount', '').strip()
    method = request.form.get('payment_method', 'bank_transfer').strip() or 'bank_transfer'
    payment_no = request.form.get('payment_no', '').strip()
    remark = request.form.get('remark', '').strip()
    attachment = ''
    if 'attachment' in request.files:
        f = request.files['attachment']
        if f and f.filename:
            _, attachment = save_upload_file(app.config['UPLOAD_FOLDER'], f, subdir='client_recharge')
    try:
        rid = record_client_recharge(
            db,
            workspace['client']['id'],
            amount=float(amount),
            payment_method=method,
            payment_no=payment_no,
            remark=remark,
            attachment=attachment,
            user_id=session.get('user_id'),
            auto_confirm=True,
        )
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '客户协同充值',
                f"公司[{workspace['company_name']}] 充值 {amount} (recharge#{rid})", request.remote_addr)
        flash('充值成功，已实时到账', 'success')
    except Exception as e:
        db.rollback()
        flash(f'充值失败：{e}', 'danger')
    return redirect(url_for('admin_client_workspace', customer_id=customer_id))


@app.route('/admin/client-company/<int:customer_id>/outbound', methods=['POST'])
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_company_outbound(customer_id):
    db = get_db()
    denied = assert_customer_access(
        db, session.get('user_id'), session.get('role'), customer_id)
    if denied:
        return denied
    item_name = request.form.get('item_name', '').strip() or '出库商品'
    quantity = request.form.get('quantity', '').strip()
    unit_price = request.form.get('unit_price', '').strip()
    amount = request.form.get('amount', '').strip()
    deduct_date = request.form.get('deduct_date', '').strip()
    remark = request.form.get('remark', '').strip()
    try:
        items = [{
            'item_name': item_name,
            'quantity': float(quantity or 0),
            'unit_price': float(unit_price or 0),
            'amount': float(amount or 0),
            'deduct_date': deduct_date,
        }]
        order_id, order_no, _, total_amount = record_client_outbound(
            db,
            customer_id=customer_id,
            client_id=None,
            items=items,
            order_date=deduct_date or None,
            remark=remark,
            user_id=session.get('user_id'),
        )
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '客户协同录入出库',
                f'order#{order_id}/{order_no} 扣减 {total_amount:.2f}', request.remote_addr)
        flash(f'出库录入成功，单号 {order_no}，扣减 {total_amount:.2f} 元', 'success')
    except Exception as e:
        db.rollback()
        flash(f'出库录入失败：{e}', 'danger')
    return redirect(url_for('admin_client_workspace', customer_id=customer_id))


@app.route('/admin/client-company/<int:customer_id>/excel-import', methods=['POST'])
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_company_excel_import(customer_id):
    db = get_db()
    denied = assert_customer_access(
        db, session.get('user_id'), session.get('role'), customer_id)
    if denied:
        return denied
    mode = request.form.get('mode', 'recharge')
    if 'excel_file' not in request.files or not request.files['excel_file'].filename:
        flash('请上传 Excel 文件', 'warning')
        return redirect(url_for('admin_client_workspace', customer_id=customer_id))
    f = request.files['excel_file']
    _, rel = save_upload_file(app.config['UPLOAD_FOLDER'], f, subdir='client_excel')
    abs_path = os.path.join(app.config['UPLOAD_FOLDER'], rel.replace('/', os.sep))
    parsed = parse_collab_excel(abs_path, mode=mode)
    if parsed.get('error'):
        flash(f'Excel 导入失败：{parsed["error"]}', 'danger')
        return redirect(url_for('admin_client_workspace', customer_id=customer_id))
    client = primary_client_for_company(db, customer_id=customer_id, client_id=None)
    if not client:
        flash('未找到激活的客户账号，无法导入', 'danger')
        return redirect(url_for('admin_client_workspace', customer_id=customer_id))
    split_orders = request.form.get('split_orders') == '1'

    if parsed.get('template') == 'standard':
        ok, msg, errs = import_standard_excel_bundle(
            db, customer_id, client['id'], parsed,
            session.get('user_id'), split_orders=split_orders,
        )
        if ok:
            db.commit()
            flash(f'标准模板导入完成：{msg}', 'success')
            if errs:
                flash(f'部分行失败：{"; ".join(errs[:5])}', 'warning')
        else:
            db.rollback()
            flash(msg, 'danger')
        return redirect(url_for('admin_client_workspace', customer_id=customer_id))

    rows = parsed['rows']
    if mode == 'recharge':
        ok, errs = import_recharge_rows(db, client['id'], rows, session.get('user_id'))
        db.commit()
        flash(f'充值导入完成：成功 {ok} 条，失败 {len(errs)} 条', 'success' if ok else 'warning')
    else:
        cnt, total, order_no, errs = import_outbound_rows(
            db, customer_id, client['id'], rows, session.get('user_id'),
            split_orders=split_orders,
        )
        if cnt > 0:
            db.commit()
            msg = f'出库导入成功：{cnt} 条明细，扣减 {total:.2f} 元'
            if order_no:
                msg += f'，单号 {order_no}'
            if errs:
                msg += f'（部分失败：{"; ".join(errs[:3])}）'
            flash(msg, 'success')
        else:
            db.rollback()
            flash(f'出库导入失败：{"; ".join(errs) if errs else "无有效数据"}', 'danger')
    return redirect(url_for('admin_client_workspace', customer_id=customer_id))


@app.route('/admin/client-company/<int:customer_id>/ocr-upload', methods=['POST'])
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_company_ocr_upload(customer_id):
    db = get_db()
    denied = assert_customer_access(
        db, session.get('user_id'), session.get('role'), customer_id)
    if denied:
        return denied
    mode = request.form.get('mode', 'recharge').strip() or 'recharge'
    if mode not in ('recharge', 'outbound'):
        mode = 'recharge'
    if 'image' not in request.files or not request.files['image'].filename:
        flash('请上传图片', 'warning')
        return redirect(url_for('admin_client_workspace', customer_id=customer_id))
    f = request.files['image']
    _, rel = save_upload_file(app.config['UPLOAD_FOLDER'], f, subdir='ocr')
    abs_path = os.path.join(app.config['UPLOAD_FOLDER'], rel.replace('/', os.sep))
    try:
        transactions, raw_text = ocr_recognize(abs_path)
        if mode == 'outbound':
            rows = ocr_rows_for_outbound(transactions, raw_text)
            label = '出库'
        else:
            rows = ocr_rows_for_recharge(transactions)
            label = '充值'
        session['collab_ocr_preview'] = {
            'customer_id': customer_id,
            'mode': mode,
            'rows': rows,
            'raw_text': raw_text[:8000],
            'image': rel,
        }
        if rows:
            flash(f'OCR 识别完成，共 {len(rows)} 条候选{label}记录，请确认后自动入账/扣减', 'success')
        else:
            flash(f'未识别到有效{label}明细，请换清晰图片或改用手工/Excel录入', 'warning')
    except Exception as e:
        flash(f'OCR 识别失败：{e}', 'danger')
    return redirect(url_for('admin_client_workspace', customer_id=customer_id))


@app.route('/admin/client-company/<int:customer_id>/ocr-apply', methods=['POST'])
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_company_ocr_apply(customer_id):
    db = get_db()
    denied = assert_customer_access(
        db, session.get('user_id'), session.get('role'), customer_id)
    if denied:
        return denied
    preview = session.get('collab_ocr_preview') or {}
    if int(preview.get('customer_id') or 0) != customer_id:
        flash('OCR 预览已过期，请重新上传图片', 'warning')
        return redirect(url_for('admin_client_workspace', customer_id=customer_id))
    rows = preview.get('rows') or []
    mode = preview.get('mode', 'recharge')
    split_orders = request.form.get('split_orders') == '1'
    if not rows:
        flash('没有可导入的 OCR 记录', 'warning')
        return redirect(url_for('admin_client_workspace', customer_id=customer_id))
    client = primary_client_for_company(db, customer_id=customer_id, client_id=None)
    if not client:
        flash('未找到激活客户账号，无法导入 OCR 结果', 'danger')
        return redirect(url_for('admin_client_workspace', customer_id=customer_id))
    if mode == 'outbound':
        cnt, total, order_no, errs = import_outbound_rows(
            db, customer_id, client['id'], rows, session.get('user_id'),
            split_orders=split_orders,
        )
        session.pop('collab_ocr_preview', None)
        if cnt > 0:
            db.commit()
            msg = f'OCR 出库导入成功：{cnt} 条明细，扣减 {total:.2f} 元'
            if order_no:
                msg += f'，单号 {order_no}'
            flash(msg, 'success')
        else:
            db.rollback()
            flash(f'OCR 出库导入失败：{"; ".join(errs) if errs else "无有效数据"}', 'danger')
    else:
        ok, errs = import_recharge_rows(db, client['id'], rows, session.get('user_id'))
        session.pop('collab_ocr_preview', None)
        db.commit()
        flash(f'OCR 充值导入完成：成功 {ok} 条，失败 {len(errs)} 条', 'success' if ok else 'warning')
    return redirect(url_for('admin_client_workspace', customer_id=customer_id))


@app.route('/admin/client-accounts')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_accounts():
    db = get_db()
    username = request.args.get('username', '').strip()
    company_name = request.args.get('company_name', '').strip()
    status = request.args.get('status', '').strip()
    sql = """SELECT ca.*, c.name as customer_master_name, u.username as approver_name
             FROM client_accounts ca
             LEFT JOIN customers c ON ca.customer_id = c.id
             LEFT JOIN users u ON ca.approved_by = u.id WHERE 1=1"""
    params = []
    if username:
        sql += " AND ca.username LIKE ?"
        params.append(f'%{username}%')
    if company_name:
        sql += " AND ca.company_name LIKE ?"
        params.append(f'%{company_name}%')
    if status:
        if status == 'approved':
            sql += " AND ca.status IN ('approved', 'active')"
        else:
            sql += " AND ca.status=?"
            params.append(status)
    scope_sql, scope_params = apply_client_account_scope_sql(
        db, session.get('user_id'), session.get('role'), 'ca')
    sql += scope_sql
    params.extend(scope_params)
    sql += " ORDER BY ca.created_at DESC"
    accounts = db.execute(sql, params).fetchall()
    total_accounts = len(accounts)
    pending_count = sum(1 for a in accounts if a['status'] == CLIENT_STATUS_PENDING)
    active_count = sum(
        1 for a in accounts
        if normalize_client_status(a['status']) == CLIENT_STATUS_APPROVED
    )
    total_balance = sum(float(a['balance'] or 0) for a in accounts)
    customers = db.execute(
        "SELECT id, name FROM customers WHERE is_active=1 ORDER BY name"
    ).fetchall()
    return render_template(
        'admin_client_accounts.html',
        accounts=accounts,
        customers=customers,
        total_accounts=total_accounts,
        pending_count=pending_count,
        active_count=active_count,
        total_balance=total_balance,
    )


@app.route('/admin/client-accounts/create', methods=['POST'])
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_accounts_create():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    phone = request.form.get('phone', '').strip()
    contact_name = request.form.get('contact_name', '').strip()
    company_name = request.form.get('company_name', '').strip()
    customer_id = request.form.get('customer_id', type=int)
    if not all([username, password, company_name]):
        flash('用户名、密码、公司名称为必填', 'warning')
        return redirect(url_for('admin_client_accounts'))
    db = get_db()
    if db.execute("SELECT id FROM client_accounts WHERE username=?", (username,)).fetchone():
        flash('用户名已存在', 'danger')
        return redirect(url_for('admin_client_accounts'))
    if customer_id:
        cust = db.execute("SELECT id FROM customers WHERE id=?", (customer_id,)).fetchone()
        if not cust:
            customer_id = None
    if not customer_id:
        customer_id = resolve_or_create_customer(db, company_name, contact_name, phone)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        """INSERT INTO client_accounts
           (customer_id, username, password, phone, contact_name, company_name,
            status, balance, total_recharge, total_deduct, approved_by, approved_at,
            created_at, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, 0, ?, ?, ?, ?)""",
        (customer_id, username, generate_password_hash(password), phone, contact_name,
         company_name, CLIENT_STATUS_APPROVED, session.get('user_id'), now, now, now),
    )
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '代建客户账号',
            f'{username} / {company_name}', request.remote_addr)
    flash('客户账号已创建并激活', 'success')
    return redirect(url_for('admin_client_accounts'))


@app.route('/admin/client-accounts/<int:id>/bind', methods=['POST'])
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_account_bind(id):
    customer_id = request.form.get('customer_id', type=int)
    db = get_db()
    account = db.execute("SELECT * FROM client_accounts WHERE id=?", (id,)).fetchone()
    if not account:
        flash('账户不存在', 'danger')
        return redirect(url_for('admin_client_accounts'))
    if not customer_id:
        flash('请选择客户主数据', 'warning')
        return redirect(url_for('admin_client_accounts'))
    cust = db.execute("SELECT name FROM customers WHERE id=?", (customer_id,)).fetchone()
    if not cust:
        flash('客户主数据不存在', 'danger')
        return redirect(url_for('admin_client_accounts'))
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        """UPDATE client_accounts SET customer_id=?, company_name=?, updated_at=? WHERE id=?""",
        (customer_id, cust['name'], now, id),
    )
    db.commit()
    flash('已绑定客户主数据', 'success')
    return redirect(url_for('admin_client_accounts'))


@app.route('/admin/client-accounts/<int:id>/reset-password', methods=['POST'])
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_account_reset_password(id):
    password = request.form.get('password', '').strip()
    if len(password) < 6:
        flash('密码至少 6 位', 'warning')
        return redirect(url_for('admin_client_accounts'))
    db = get_db()
    db.execute(
        "UPDATE client_accounts SET password=?, updated_at=? WHERE id=?",
        (generate_password_hash(password),
         datetime.now().strftime('%Y-%m-%d %H:%M:%S'), id),
    )
    db.commit()
    flash('密码已重置', 'success')
    return redirect(url_for('admin_client_accounts'))


@app.route('/admin/client-accounts/<int:id>/approve')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_account_approve(id):
    db = get_db()
    account = db.execute("SELECT * FROM client_accounts WHERE id=?", (id,)).fetchone()
    if not account:
        flash('账户不存在', 'danger')
        return redirect(url_for('admin_client_accounts'))
    if account['status'] != CLIENT_STATUS_PENDING:
        flash('该账户不在待审核状态', 'warning')
        return redirect(url_for('admin_client_accounts'))
    customer_id = account['customer_id']
    if not customer_id:
        customer_id = resolve_or_create_customer(
            db, account['company_name'], account['contact_name'], account['phone'])
        db.execute(
            "UPDATE client_accounts SET customer_id=? WHERE id=?",
            (customer_id, id),
        )
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        """UPDATE client_accounts SET status=?, approved_by=?, approved_at=?, updated_at=?
           WHERE id=?""",
        (CLIENT_STATUS_APPROVED, session.get('user_id'), now, now, id),
    )
    if customer_id and session.get('role') == ROLE_CLIENT_COLLAB:
        assign_customer_to_user(
            db, session.get('user_id'), customer_id, session.get('user_id'))
    db.execute(
        """INSERT INTO client_messages (client_id, title, content, msg_type)
           VALUES (?, ?, ?, 'notice')""",
        (id, '账户审核通过', '您的账户已审核通过，可以正常登录使用系统。'),
    )
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '审核客户账户',
            f"通过 {account['username']}", request.remote_addr)
    flash('已审核通过', 'success')
    return redirect(url_for('admin_client_accounts'))


@app.route('/admin/client-accounts/<int:id>/reject')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_account_reject(id):
    db = get_db()
    account = db.execute("SELECT * FROM client_accounts WHERE id=?", (id,)).fetchone()
    if not account:
        flash('账户不存在', 'danger')
        return redirect(url_for('admin_client_accounts'))
    if account['status'] != CLIENT_STATUS_PENDING:
        flash('该账户不在待审核状态', 'warning')
        return redirect(url_for('admin_client_accounts'))
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        """UPDATE client_accounts SET status=?, approved_by=?, approved_at=?, updated_at=?
           WHERE id=?""",
        (CLIENT_STATUS_REJECTED, session.get('user_id'), now, now, id),
    )
    db.execute(
        """INSERT INTO client_messages (client_id, title, content, msg_type)
           VALUES (?, ?, ?, 'notice')""",
        (id, '账户审核未通过', '您的账户审核未通过，请联系管理员了解详情。'),
    )
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '审核客户账户',
            f"拒绝 {account['username']}", request.remote_addr)
    flash('已拒绝该账户', 'success')
    return redirect(url_for('admin_client_accounts'))


@app.route('/admin/client-recharges')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_recharges():
    db = get_db()
    client_id = request.args.get('client_id', type=int)
    status = request.args.get('status', '').strip()
    payment_method = request.args.get('payment_method', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    sql = """SELECT cr.*, ca.username, ca.company_name, ca.contact_name,
                    u.username as confirmer_name
             FROM client_recharges cr
             LEFT JOIN client_accounts ca ON cr.client_id = ca.id
             LEFT JOIN users u ON cr.confirmed_by = u.id WHERE 1=1"""
    params = []
    if client_id:
        sql += " AND cr.client_id=?"
        params.append(client_id)
    if status:
        sql += " AND cr.status=?"
        params.append(status)
    if payment_method:
        sql += " AND cr.payment_method=?"
        params.append(payment_method)
    if start_date:
        sql += " AND date(cr.created_at) >= ?"
        params.append(start_date)
    if end_date:
        sql += " AND date(cr.created_at) <= ?"
        params.append(end_date)
    scope_sql, scope_params = apply_client_account_scope_sql(
        db, session.get('user_id'), session.get('role'), 'ca')
    sql += scope_sql
    params.extend(scope_params)
    sql += " ORDER BY cr.created_at DESC"
    recharges = db.execute(sql, params).fetchall()
    clients_sql = "SELECT id, company_name FROM client_accounts ca WHERE 1=1"
    clients_params = []
    cscope, cscope_p = apply_client_account_scope_sql(
        db, session.get('user_id'), session.get('role'), 'ca')
    clients_sql += cscope
    clients_params.extend(cscope_p)
    clients_sql += " ORDER BY company_name"
    clients = db.execute(clients_sql, clients_params).fetchall()
    return render_template('admin_client_recharges.html', recharges=recharges, clients=clients)


@app.route('/admin/client-recharges/<int:id>/confirm')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_recharge_confirm(id):
    db = get_db()
    recharge = db.execute("SELECT * FROM client_recharges WHERE id=?", (id,)).fetchone()
    if not recharge:
        flash('充值记录不存在', 'danger')
        return redirect(url_for('admin_client_recharges'))
    if recharge['status'] != 'pending':
        flash('该充值记录不在待确认状态', 'warning')
        return redirect(url_for('admin_client_recharges'))
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "UPDATE client_recharges SET status='confirmed', confirmed_by=?, confirmed_at=? WHERE id=?",
        (session.get('user_id'), now, id),
    )
    db.execute(
        """UPDATE client_accounts SET balance=balance+?, total_recharge=total_recharge+?, updated_at=?
           WHERE id=?""",
        (recharge['amount'], recharge['amount'], now, recharge['client_id']),
    )
    db.execute(
        """INSERT INTO client_messages (client_id, title, content, msg_type)
           VALUES (?, ?, ?, 'notice')""",
        (recharge['client_id'], '充值成功',
         f'您的充值申请 {recharge["amount"]:.2f} 元已确认到账。'),
    )
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '确认充值',
            f"充值单 {id} 金额 {recharge['amount']:.2f}", request.remote_addr)
    flash('充值已确认', 'success')
    return redirect(url_for('admin_client_recharges'))


@app.route('/admin/client-recharges/<int:id>/reject')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_recharge_reject(id):
    db = get_db()
    recharge = db.execute("SELECT * FROM client_recharges WHERE id=?", (id,)).fetchone()
    if not recharge:
        flash('充值记录不存在', 'danger')
        return redirect(url_for('admin_client_recharges'))
    if recharge['status'] != 'pending':
        flash('该充值记录不在待确认状态', 'warning')
        return redirect(url_for('admin_client_recharges'))
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "UPDATE client_recharges SET status='rejected', confirmed_by=?, confirmed_at=? WHERE id=?",
        (session.get('user_id'), now, id),
    )
    db.execute(
        """INSERT INTO client_messages (client_id, title, content, msg_type)
           VALUES (?, ?, ?, 'notice')""",
        (recharge['client_id'], '充值申请被拒绝',
         f'您的充值申请 {recharge["amount"]:.2f} 元已被拒绝，请联系管理员了解详情。'),
    )
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '拒绝充值',
            f"充值单 {id}", request.remote_addr)
    flash('已拒绝该充值申请', 'success')
    return redirect(url_for('admin_client_recharges'))


@app.route('/admin/client-deductions/sync')
@login_required
@module_required(MODULE_CLIENT_PORTAL)
def admin_client_deductions_sync():
    db = get_db()
    count = sync_deductions_for_customer(db)
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '同步扣减',
            f"同步 {count} 条", request.remote_addr)
    if count == 0:
        flash('没有需要同步的出库记录', 'info')
    else:
        flash(f'同步完成，共处理 {count} 条扣减记录', 'success')
    return redirect(url_for('admin_client_accounts'))


# ==================== 采购管理 ====================

def _table_columns(db, table):
    return {r[1] for r in db.execute(f'PRAGMA table_info({table})').fetchall()}


def _next_purchase_no(db):
    prefix = f'PO{datetime.now().strftime("%Y%m%d")}'
    rows = db.execute(
        "SELECT purchase_no FROM purchase_orders WHERE purchase_no LIKE ?",
        (f'{prefix}%',),
    ).fetchall()
    max_seq = 0
    for row in rows:
        no = row['purchase_no'] or ''
        suffix = no.replace(prefix, '', 1)
        if suffix.isdigit():
            max_seq = max(max_seq, int(suffix))
    return f'{prefix}{str(max_seq + 1).zfill(4)}'


def _purchase_no_exists(db, purchase_no, exclude_id=None):
    if not purchase_no:
        return False
    sql = "SELECT id FROM purchase_orders WHERE purchase_no=?"
    params = [purchase_no]
    if exclude_id:
        sql += " AND id<>?"
        params.append(exclude_id)
    return db.execute(sql, params).fetchone() is not None


def _save_purchase_attachment(file_storage):
    if not file_storage or not file_storage.filename:
        return ''
    ext = os.path.splitext(file_storage.filename)[1]
    filename = f'purchase_{uuid.uuid4().hex[:8]}{ext}'
    folder = os.path.join(app.config['UPLOAD_FOLDER'], 'purchases')
    os.makedirs(folder, exist_ok=True)
    file_storage.save(os.path.join(folder, filename))
    return filename


def _purchase_form_payload(db, existing=None):
    cols = _table_columns(db, 'purchase_orders')
    payload = {}

    def set_if_col(column, value):
        if column in cols:
            payload[column] = value

    purchase_no = request.form.get('purchase_no', '').strip()
    if not purchase_no:
        purchase_no = _next_purchase_no(db)

    total_amount = request.form.get('total_amount', type=float)
    if total_amount is None:
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('unit_price[]')
        total_amount = 0
        for q, p in zip(quantities, prices):
            try:
                total_amount += float(q or 0) * float(p or 0)
            except ValueError:
                pass

    set_if_col('project_id', request.form.get('project_id', type=int))
    set_if_col('contract_id', request.form.get('contract_id', type=int) or None)
    set_if_col('purchase_no', purchase_no)
    set_if_col('purchase_type', request.form.get('purchase_type', '材料采购'))
    set_if_col('supplier', request.form.get('supplier', '').strip())
    set_if_col('description', request.form.get('description', '').strip())
    set_if_col('total_amount', total_amount or 0)
    purchase_date = request.form.get('purchase_date', '') or request.form.get('order_date', '')
    set_if_col('purchase_date', purchase_date)
    set_if_col('order_date', purchase_date)
    set_if_col('delivery_address', request.form.get('delivery_address', '').strip())
    set_if_col('status', request.form.get('status') or (existing['status'] if existing and 'status' in existing.keys() else 'draft'))
    set_if_col('remark', request.form.get('remark', '').strip())

    attachment = existing['attachment'] if existing and 'attachment' in existing.keys() else ''
    if 'attachment' in request.files:
        uploaded = _save_purchase_attachment(request.files['attachment'])
        if uploaded:
            attachment = uploaded
    set_if_col('attachment', attachment)

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if existing is None:
        set_if_col('created_by', session.get('user_id'))
        set_if_col('created_at', now)
    set_if_col('updated_at', now)
    return payload


def _replace_purchase_items(db, purchase_id):
    item_names = request.form.getlist('item_name[]')
    specs = request.form.getlist('specification[]')
    units = request.form.getlist('unit[]')
    quantities = request.form.getlist('quantity[]')
    prices = request.form.getlist('unit_price[]')

    db.execute("DELETE FROM purchase_items WHERE purchase_id=?", (purchase_id,))
    item_cols = _table_columns(db, 'purchase_items')
    for idx, name in enumerate(item_names):
        name = (name or '').strip()
        if not name:
            continue
        try:
            qty = float(quantities[idx] or 0) if idx < len(quantities) else 0
        except ValueError:
            qty = 0
        try:
            price = float(prices[idx] or 0) if idx < len(prices) else 0
        except ValueError:
            price = 0
        row = {
            'purchase_id': purchase_id,
            'item_name': name,
            'specification': specs[idx].strip() if idx < len(specs) else '',
            'unit': units[idx] if idx < len(units) else '吨',
            'quantity': qty,
            'unit_price': price,
            'amount': round(qty * price, 2),
            'sort_order': idx,
        }
        row = {k: v for k, v in row.items() if k in item_cols}
        columns = ', '.join(row.keys())
        placeholders = ', '.join('?' for _ in row)
        db.execute(
            f"INSERT INTO purchase_items ({columns}) VALUES ({placeholders})",
            list(row.values()),
        )


@app.route('/purchase/list')
@login_required
def purchase_list():
    """采购单列表"""
    db = get_db()
    project_id = request.args.get('project_id', type=int)
    status = request.args.get('status', '')
    keyword = request.args.get('keyword', '')

    sql = """SELECT po.*, p.name as project_name, c.contract_name,
                    (SELECT COUNT(DISTINCT tpi.transport_id) 
                     FROM transport_purchase_items tpi 
                     JOIN purchase_items pi ON tpi.purchase_item_id = pi.id 
                     WHERE pi.purchase_id = po.id) as transport_count
             FROM purchase_orders po
             LEFT JOIN projects p ON po.project_id = p.id
             LEFT JOIN contracts c ON po.contract_id = c.id
             WHERE 1=1"""
    params = []

    if project_id:
        sql += " AND po.project_id = ?"
        params.append(project_id)
    if status:
        sql += " AND po.status = ?"
        params.append(status)
    if keyword:
        sql += " AND (po.purchase_no LIKE ? OR po.supplier LIKE ?)"
        params.extend([f'%{keyword}%', f'%{keyword}%'])

    sql += " ORDER BY po.created_at DESC"
    purchases = db.execute(sql, params).fetchall()

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    filters = {'project_id': project_id, 'status': status, 'keyword': keyword}
    return render_template('purchase_list.html', purchases=purchases, projects=projects,
                           filters=filters)


@app.route('/purchase/add', methods=['GET', 'POST'])
@login_required
def purchase_add():
    """新增采购单"""
    db = get_db()
    if request.method == 'POST':
        payload = _purchase_form_payload(db)
        original_purchase_no = payload.get('purchase_no')
        if _purchase_no_exists(db, original_purchase_no):
            payload['purchase_no'] = _next_purchase_no(db)
            flash(f'采购单号 {original_purchase_no} 已存在，系统已自动改为 {payload["purchase_no"]}', 'warning')
        columns = ', '.join(payload.keys())
        placeholders = ', '.join('?' for _ in payload)
        try:
            cur = db.execute(
                f"INSERT INTO purchase_orders ({columns}) VALUES ({placeholders})",
                list(payload.values()),
            )
        except sqlite3.IntegrityError as e:
            if 'purchase_orders.purchase_no' in str(e):
                payload['purchase_no'] = _next_purchase_no(db)
                cur = db.execute(
                    f"INSERT INTO purchase_orders ({columns}) VALUES ({placeholders})",
                    list(payload.values()),
                )
                flash(f'采购单号重复，系统已自动改为 {payload["purchase_no"]}', 'warning')
            else:
                raise
        purchase_id = cur.lastrowid
        _replace_purchase_items(db, purchase_id)
        db.commit()
        purchase_no = payload.get('purchase_no') or purchase_id
        add_log(session.get('user_id'), session.get('username', ''), '新增采购单', f'采购单号: {purchase_no}')
        flash('采购单创建成功', 'success')
        return redirect(url_for('purchase_detail', id=purchase_id))

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    contracts = db.execute("SELECT id, contract_name, contract_type FROM contracts ORDER BY contract_name").fetchall()
    suppliers = db.execute("SELECT id, name FROM suppliers WHERE is_active=1 ORDER BY name").fetchall()
    next_purchase_no = _next_purchase_no(db)
    return render_template('purchase_form.html', projects=projects, contracts=contracts,
                           suppliers=suppliers, now=datetime.now(), next_purchase_no=next_purchase_no)


@app.route('/purchase/<int:id>')
@login_required
def purchase_detail(id):
    """采购单详情"""
    db = get_db()
    purchase = db.execute("""SELECT po.*, p.name as project_name, c.contract_name
                             FROM purchase_orders po
                             LEFT JOIN projects p ON po.project_id = p.id
                             LEFT JOIN contracts c ON po.contract_id = c.id
                             WHERE po.id = ?""", (id,)).fetchone()
    if not purchase:
        flash('采购单不存在', 'danger')
        return redirect(url_for('purchase_list'))

    purchase = dict(purchase)
    if not purchase.get('purchase_date') and purchase.get('order_date'):
        purchase['purchase_date'] = purchase['order_date']

    items = db.execute("""SELECT pi.*,
                                 (SELECT COALESCE(SUM(tpi.quantity), 0)
                                  FROM transport_purchase_items tpi
                                  WHERE tpi.purchase_item_id = pi.id) as linked_quantity
                          FROM purchase_items pi
                          WHERE pi.purchase_id = ?
                          ORDER BY pi.sort_order, pi.id""", (id,)).fetchall()

    total_purchase_qty = sum(float(i['quantity'] or 0) for i in items)

    tr_cols = {r[1] for r in db.execute('PRAGMA table_info(transport_records)').fetchall()}
    inv_cols = {r[1] for r in db.execute('PRAGMA table_info(invoices)').fetchall()}
    inv_join = ' LEFT JOIN invoices i ON tr.invoice_id = i.id' if 'invoice_id' in tr_cols and inv_cols else ''
    inv_sel = ', i.invoice_no' if 'invoice_id' in tr_cols and inv_cols else ", NULL as invoice_no"

    if 'purchase_id' in tr_cols:
        transports = db.execute(f"""
            SELECT tr.*{inv_sel}
            FROM transport_records tr{inv_join}
            WHERE tr.purchase_id = ?
            ORDER BY tr.transport_date, tr.id
        """, (id,)).fetchall()
    else:
        transport_ids = db.execute("""SELECT DISTINCT tpi.transport_id
                                      FROM transport_purchase_items tpi
                                      JOIN purchase_items pi ON tpi.purchase_item_id = pi.id
                                      WHERE pi.purchase_id = ?""", (id,)).fetchall()
        transports = []
        if transport_ids:
            tid_str = ','.join(str(t['transport_id']) for t in transport_ids)
            transports = db.execute(f"""
                SELECT tr.*{inv_sel}
                FROM transport_records tr{inv_join}
                WHERE tr.id IN ({tid_str})
                ORDER BY tr.transport_date, tr.id
            """).fetchall()

    total_transport_qty = sum(float(t['quantity'] or 0) for t in transports)
    total_freight = sum(float(t['freight_amount'] or 0) for t in transports)

    transport_item_names = {}
    for t in transports:
        tid = t['id']
        names = db.execute("""
            SELECT pi.item_name FROM transport_purchase_items tpi
            JOIN purchase_items pi ON tpi.purchase_item_id = pi.id
            WHERE tpi.transport_id = ?
        """, (tid,)).fetchall()
        if names:
            transport_item_names[tid] = '、'.join(r['item_name'] for r in names)
        elif 'purchase_item_id' in tr_cols and t['purchase_item_id']:
            item_row = db.execute(
                "SELECT item_name FROM purchase_items WHERE id=?", (t['purchase_item_id'],)
            ).fetchone()
            if item_row:
                transport_item_names[tid] = item_row['item_name']

    reconciliation = None
    recon_cols = {r[1] for r in db.execute('PRAGMA table_info(reconciliations)').fetchall()}
    if 'purchase_id' in recon_cols:
        reconciliation = db.execute(
            "SELECT * FROM reconciliations WHERE purchase_id=? ORDER BY id DESC LIMIT 1",
            (id,),
        ).fetchone()

    if purchase.get('loss_rate') is None:
        if reconciliation and reconciliation['allowed_loss_rate'] is not None:
            purchase['loss_rate'] = reconciliation['allowed_loss_rate']
        else:
            purchase['loss_rate'] = 0.5

    return render_template(
        'purchase_detail.html',
        purchase=purchase,
        items=items,
        transports=transports,
        transport_item_names=transport_item_names,
        total_purchase_qty=total_purchase_qty,
        total_transport_qty=total_transport_qty,
        total_freight=total_freight,
        reconciliation=reconciliation,
    )


@app.route('/purchase/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def purchase_edit(id):
    """编辑采购单"""
    db = get_db()
    purchase = db.execute("SELECT * FROM purchase_orders WHERE id = ?", (id,)).fetchone()
    if not purchase:
        flash('采购单不存在', 'danger')
        return redirect(url_for('purchase_list'))

    if request.method == 'POST':
        payload = _purchase_form_payload(db, purchase)
        if _purchase_no_exists(db, payload.get('purchase_no'), exclude_id=id):
            flash(f'采购单号 {payload.get("purchase_no")} 已存在，请换一个单号', 'danger')
            return redirect(url_for('purchase_edit', id=id))
        assignments = ', '.join(f'{k}=?' for k in payload.keys())
        try:
            db.execute(
                f"UPDATE purchase_orders SET {assignments} WHERE id=?",
                list(payload.values()) + [id],
            )
        except sqlite3.IntegrityError as e:
            if 'purchase_orders.purchase_no' in str(e):
                flash('采购单号已存在，请换一个单号', 'danger')
                return redirect(url_for('purchase_edit', id=id))
            raise
        _replace_purchase_items(db, id)
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '编辑采购单', f'采购单ID: {id}')
        flash('采购单更新成功', 'success')
        return redirect(url_for('purchase_detail', id=id))

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    contracts = db.execute("SELECT id, contract_name, contract_type FROM contracts ORDER BY contract_name").fetchall()
    suppliers = db.execute("SELECT id, name FROM suppliers WHERE is_active=1 ORDER BY name").fetchall()
    items = db.execute(
        "SELECT * FROM purchase_items WHERE purchase_id=? ORDER BY sort_order, id",
        (id,),
    ).fetchall()
    return render_template('purchase_form.html', purchase=purchase, projects=projects,
                           contracts=contracts, suppliers=suppliers, items=items)


@app.route('/purchase/<int:id>/delete', methods=['POST'])
@login_required
def purchase_delete(id):
    """删除采购单"""
    db = get_db()
    purchase = db.execute("SELECT * FROM purchase_orders WHERE id = ?", (id,)).fetchone()
    if not purchase:
        flash('采购单不存在', 'danger')
        return redirect(url_for('purchase_list'))

    # 删除关联数据
    db.execute("DELETE FROM transport_purchase_items WHERE purchase_item_id IN (SELECT id FROM purchase_items WHERE purchase_id=?)", (id,))
    db.execute("DELETE FROM purchase_items WHERE purchase_id = ?", (id,))
    db.execute("DELETE FROM purchase_orders WHERE id = ?", (id,))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '删除采购单', f'采购单ID: {id}')
    flash('采购单已删除', 'success')
    return redirect(url_for('purchase_list'))


def _coerce_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


@app.route('/api/purchase/<int:purchase_id>/item/add', methods=['POST'])
@app.route('/api/purchase/item/add', methods=['POST'])
@login_required
def api_purchase_item_add(purchase_id=None):
    """API添加采购明细"""
    db = get_db()
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        payload = request.form.to_dict()

    if purchase_id is None:
        purchase_id = payload.get('purchase_id')
    try:
        purchase_id = int(purchase_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '缺少采购单ID'}), 400

    order = db.execute("SELECT id FROM purchase_orders WHERE id=?", (purchase_id,)).fetchone()
    if not order:
        return jsonify({'success': False, 'error': '采购单不存在'}), 404

    item_name = (payload.get('item_name') or '').strip()
    if not item_name:
        return jsonify({'success': False, 'error': '请输入品名'}), 400
    specification = (payload.get('specification') or '').strip()
    unit = (payload.get('unit') or '吨').strip()
    quantity = _coerce_float(payload.get('quantity'))
    unit_price = _coerce_float(payload.get('unit_price'))
    amount = round(quantity * unit_price, 2)

    max_order = db.execute(
        "SELECT COALESCE(MAX(sort_order), 0) FROM purchase_items WHERE purchase_id=?",
        (purchase_id,),
    ).fetchone()[0]

    cur = db.execute(
        """INSERT INTO purchase_items
           (purchase_id, item_name, specification, unit, quantity, unit_price, amount, sort_order)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (purchase_id, item_name, specification, unit, quantity, unit_price, amount, max_order + 1),
    )
    item_id = cur.lastrowid

    total = db.execute(
        "SELECT COALESCE(SUM(amount), 0) FROM purchase_items WHERE purchase_id=?",
        (purchase_id,),
    ).fetchone()[0]
    db.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?", (total, purchase_id))
    db.commit()

    return jsonify({'success': True, 'item_id': item_id, 'amount': amount, 'total': total})


@app.route('/api/purchase/item/<int:item_id>/edit', methods=['POST'])
@login_required
def api_purchase_item_edit(item_id):
    """API编辑采购明细"""
    db = get_db()
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        payload = request.form.to_dict()
    item = db.execute("SELECT * FROM purchase_items WHERE id=?", (item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': '明细不存在'}), 404

    item_name = (payload.get('item_name') or item['item_name'] or '').strip()
    specification = payload.get('specification', item['specification'])
    unit = payload.get('unit', item['unit']) or '吨'
    quantity = _coerce_float(payload.get('quantity'), item['quantity'] or 0)
    unit_price = _coerce_float(payload.get('unit_price'), item['unit_price'] or 0)
    amount = round(quantity * unit_price, 2)

    db.execute("""UPDATE purchase_items SET item_name=?, specification=?, unit=?, quantity=?, unit_price=?, amount=?
                  WHERE id=?""", (item_name, specification, unit, quantity, unit_price, amount, item_id))

    total = db.execute("SELECT COALESCE(SUM(amount), 0) FROM purchase_items WHERE purchase_id=?", (item['purchase_id'],)).fetchone()[0]
    db.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?", (total, item['purchase_id']))
    db.commit()
    return jsonify({'success': True, 'amount': amount, 'total': total})


@app.route('/api/purchase/item/<int:item_id>/delete', methods=['POST'])
@login_required
def api_purchase_item_delete(item_id):
    """API删除采购明细"""
    db = get_db()
    item = db.execute("SELECT * FROM purchase_items WHERE id=?", (item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': '明细不存在'}), 404

    purchase_id = item['purchase_id']
    db.execute("DELETE FROM transport_purchase_items WHERE purchase_item_id=?", (item_id,))
    db.execute("DELETE FROM purchase_items WHERE id=?", (item_id,))

    total = db.execute("SELECT COALESCE(SUM(amount), 0) FROM purchase_items WHERE purchase_id=?", (purchase_id,)).fetchone()[0]
    db.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?", (total, purchase_id))
    db.commit()
    return jsonify({'success': True, 'total': total})


@app.route('/purchase/<int:purchase_id>/item/import', methods=['GET', 'POST'])
@login_required
def purchase_item_import(purchase_id):
    """采购明细导入"""
    db = get_db()
    purchase = db.execute("SELECT * FROM purchase_orders WHERE id=?", (purchase_id,)).fetchone()
    if not purchase:
        flash('采购单不存在', 'danger')
        return redirect(url_for('purchase_list'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择文件', 'danger')
            return redirect(request.url)

        f = request.files['file']
        if not f.filename:
            flash('请选择文件', 'danger')
            return redirect(request.url)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            count = 0
            max_order = db.execute("SELECT COALESCE(MAX(sort_order), 0) FROM purchase_items WHERE purchase_id=?", (purchase_id,)).fetchone()[0]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                item_name = str(row[0]).strip() if row[0] else ''
                specification = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                unit = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                quantity = float(row[3]) if len(row) > 3 and row[3] else 0
                unit_price = float(row[4]) if len(row) > 4 and row[4] else 0
                amount = round(quantity * unit_price, 2)
                max_order += 1

                db.execute("""INSERT INTO purchase_items (purchase_id, item_name, specification, unit, quantity, unit_price, amount, sort_order)
                              VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                           (purchase_id, item_name, specification, unit, quantity, unit_price, amount, max_order))
                count += 1

            total = db.execute("SELECT COALESCE(SUM(amount), 0) FROM purchase_items WHERE purchase_id=?", (purchase_id,)).fetchone()[0]
            db.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?", (total, purchase_id))
            db.commit()
            add_log(session.get('user_id'), session.get('username', ''), '导入采购明细', f'采购单ID: {purchase_id}, 导入{count}条')
            flash(f'成功导入 {count} 条明细', 'success')
        except Exception as e:
            flash(f'导入失败: {str(e)}', 'danger')

        return redirect(url_for('purchase_detail', id=purchase_id))

    return render_template('purchase_item_import.html', purchase=purchase)


# ==================== 销售管理 ====================

@app.route('/sales/contract/list')
@login_required
def sales_contract_list():
    """销售合同列表"""
    db = get_db()
    contracts = db.execute("""SELECT c.*, p.name as project_name
                              FROM contracts c
                              LEFT JOIN projects p ON c.project_id = p.id
                              WHERE c.contract_type IN ('销售合同', '收入合同')
                              ORDER BY c.created_at DESC""").fetchall()
    return render_template('sales_contract_list.html', contracts=contracts)


@app.route('/sales/payment/list')
@login_required
def sales_payment_list():
    """销售回款列表"""
    db = get_db()
    project_id = request.args.get('project_id', type=int)
    customer = request.args.get('customer', '')

    sql = """SELECT sp.*, p.name as project_name, c.contract_name
             FROM sales_payments sp
             LEFT JOIN projects p ON sp.project_id = p.id
             LEFT JOIN contracts c ON sp.contract_id = c.id
             WHERE 1=1"""
    params = []
    if project_id:
        sql += " AND sp.project_id = ?"
        params.append(project_id)
    if customer:
        sql += " AND sp.customer_name LIKE ?"
        params.append(f'%{customer}%')
    sql += " ORDER BY sp.created_at DESC"

    payments = db.execute(sql, params).fetchall()
    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    filters = {'project_id': project_id, 'customer': customer}
    return render_template('sales_payment_list.html', payments=payments, projects=projects,
                           filters=filters)


@app.route('/sales/payment/add', methods=['GET', 'POST'])
@login_required
def sales_payment_add():
    """新增销售回款"""
    db = get_db()
    if request.method == 'POST':
        project_id = request.form.get('project_id', type=int)
        contract_id = request.form.get('contract_id', type=int) or None
        customer_name = request.form.get('customer_name', '')
        payment_no = request.form.get('payment_no', '')
        amount = request.form.get('amount', type=float) or 0
        payment_date = request.form.get('payment_date', '')
        payment_method = request.form.get('payment_method', '')
        remark = request.form.get('remark', '')

        db.execute("""INSERT INTO sales_payments (project_id, contract_id, customer_name, payment_no,
                      amount, payment_date, payment_method, remark, created_by, created_at)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                   (project_id, contract_id, customer_name, payment_no, amount, payment_date,
                    payment_method, remark, session.get('user_id'), datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '新增销售回款', f'客户: {customer_name}, 金额: {amount}')
        flash('销售回款添加成功', 'success')
        return redirect(url_for('sales_payment_list'))

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    contracts = db.execute("SELECT id, contract_name FROM contracts WHERE contract_type IN ('销售合同', '收入合同') ORDER BY contract_name").fetchall()
    customers = db.execute("SELECT id, name FROM customers WHERE is_active=1 ORDER BY name").fetchall()
    return render_template('sales_payment_form.html', projects=projects, contracts=contracts,
                           customers=customers, now=datetime.now())


@app.route('/sales/order/list')
@login_required
def sales_order_list():
    """销售出库单列表"""
    db = get_db()
    project_id = request.args.get('project_id', type=int)
    status = request.args.get('status', '')
    keyword = request.args.get('keyword', '')

    sql = """SELECT so.*, p.name as project_name, c.contract_name
             FROM sales_orders so
             LEFT JOIN projects p ON so.project_id = p.id
             LEFT JOIN contracts c ON so.contract_id = c.id
             WHERE 1=1"""
    params = []

    if project_id:
        sql += " AND so.project_id = ?"
        params.append(project_id)
    if status:
        sql += " AND so.status = ?"
        params.append(status)
    if keyword:
        sql += " AND (so.order_no LIKE ? OR so.customer_name LIKE ?)"
        params.extend([f'%{keyword}%', f'%{keyword}%'])

    sql += " ORDER BY so.created_at DESC"
    orders = db.execute(sql, params).fetchall()
    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    filters = {'project_id': project_id, 'status': status, 'keyword': keyword}
    return render_template('sales_order_list.html', orders=orders, projects=projects,
                           filters=filters)


@app.route('/sales/order/add', methods=['GET', 'POST'])
@login_required
def sales_order_add():
    """新增销售出库单"""
    db = get_db()
    if request.method == 'POST':
        project_id = request.form.get('project_id', type=int)
        contract_id = request.form.get('contract_id', type=int) or None
        customer_name = request.form.get('customer_name', '')
        order_date = request.form.get('order_date', '')
        delivery_date = request.form.get('delivery_date', '') or None
        remark = request.form.get('remark', '')

        # 生成出库单号
        count = db.execute("SELECT COUNT(*) FROM sales_orders").fetchone()[0]
        order_no = f'SO{datetime.now().strftime("%Y%m%d")}{str(count + 1).zfill(4)}'

        # 处理附件
        attachment = ''
        if 'attachment' in request.files:
            f = request.files['attachment']
            if f and f.filename:
                ext = os.path.splitext(f.filename)[1]
                filename = f'sales_{uuid.uuid4().hex[:8]}{ext}'
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                attachment = filename

        customer_id = request.form.get('customer_id', type=int)
        if not customer_id and customer_name:
            crow = db.execute(
                "SELECT id FROM customers WHERE name=?", (customer_name,)
            ).fetchone()
            if crow:
                customer_id = crow['id']
        db.execute("""INSERT INTO sales_orders (project_id, contract_id, order_no, customer_name,
                      customer_id, order_date, delivery_date, total_amount, total_quantity, status,
                      remark, attachment, created_by, created_at)
                      VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, '待审核', ?, ?, ?, ?)""",
                   (project_id, contract_id, order_no, customer_name, customer_id,
                    order_date, delivery_date, remark, attachment, session.get('user_id'),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '新增销售出库单', f'出库单号: {order_no}')
        flash('销售出库单创建成功', 'success')
        return redirect(url_for('sales_order_list'))

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    contracts = db.execute("SELECT id, contract_name FROM contracts WHERE contract_type IN ('销售合同', '收入合同') ORDER BY contract_name").fetchall()
    customers = db.execute("SELECT id, name FROM customers WHERE is_active=1 ORDER BY name").fetchall()
    return render_template('sales_order_form.html', projects=projects, contracts=contracts,
                           customers=customers, now=datetime.now())


@app.route('/sales/order/<int:id>')
@login_required
def sales_order_detail(id):
    """销售出库单详情"""
    db = get_db()
    order = db.execute("""SELECT so.*, p.name as project_name, c.contract_name
                          FROM sales_orders so
                          LEFT JOIN projects p ON so.project_id = p.id
                          LEFT JOIN contracts c ON so.contract_id = c.id
                          WHERE so.id = ?""", (id,)).fetchone()
    if not order:
        flash('出库单不存在', 'danger')
        return redirect(url_for('sales_order_list'))

    items = db.execute("""SELECT soi.*,
                                 (SELECT COALESCE(SUM(sit.quantity), 0) FROM sales_item_transport sit WHERE sit.sales_item_id = soi.id) as linked_quantity
                          FROM sales_order_items soi
                          WHERE soi.sales_order_id = ?
                          ORDER BY soi.sort_order, soi.id""", (id,)).fetchall()

    # 关联的运输记录
    transport_ids = db.execute("""SELECT DISTINCT sit.transport_id
                                  FROM sales_item_transport sit
                                  JOIN sales_order_items soi ON sit.sales_item_id = soi.id
                                  WHERE soi.sales_order_id = ?""", (id,)).fetchall()
    transport_list = []
    if transport_ids:
        tid_str = ','.join([str(t['transport_id']) for t in transport_ids])
        transport_list = db.execute(f"""SELECT tr.* FROM transport_records tr
                                         WHERE tr.id IN ({tid_str}) ORDER BY tr.transport_date""").fetchall()

    return render_template('sales_order_detail.html', order=order, items=items,
                           transport_list=transport_list)


@app.route('/sales/order/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def sales_order_edit(id):
    """编辑销售出库单"""
    db = get_db()
    order = db.execute("SELECT * FROM sales_orders WHERE id = ?", (id,)).fetchone()
    if not order:
        flash('出库单不存在', 'danger')
        return redirect(url_for('sales_order_list'))

    if request.method == 'POST':
        project_id = request.form.get('project_id', type=int)
        contract_id = request.form.get('contract_id', type=int) or None
        customer_name = request.form.get('customer_name', '')
        order_date = request.form.get('order_date', '')
        delivery_date = request.form.get('delivery_date', '') or None
        status = request.form.get('status', '待审核')
        remark = request.form.get('remark', '')

        attachment = order['attachment']
        if 'attachment' in request.files:
            f = request.files['attachment']
            if f and f.filename:
                ext = os.path.splitext(f.filename)[1]
                filename = f'sales_{uuid.uuid4().hex[:8]}{ext}'
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                attachment = filename

        db.execute("""UPDATE sales_orders SET project_id=?, contract_id=?, customer_name=?,
                      order_date=?, delivery_date=?, status=?, remark=?, attachment=?, updated_at=?
                      WHERE id=?""",
                   (project_id, contract_id, customer_name, order_date, delivery_date, status,
                    remark, attachment, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), id))
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '编辑销售出库单', f'出库单ID: {id}')
        flash('出库单更新成功', 'success')
        return redirect(url_for('sales_order_detail', id=id))

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    contracts = db.execute("SELECT id, contract_name FROM contracts WHERE contract_type IN ('销售合同', '收入合同') ORDER BY contract_name").fetchall()
    customers = db.execute("SELECT id, name FROM customers WHERE is_active=1 ORDER BY name").fetchall()
    return render_template('sales_order_form.html', order=order, projects=projects, edit_mode=True,
                           contracts=contracts, customers=customers)


@app.route('/sales/order/<int:id>/delete', methods=['POST'])
@login_required
def sales_order_delete(id):
    """删除销售出库单"""
    db = get_db()
    order = db.execute("SELECT * FROM sales_orders WHERE id = ?", (id,)).fetchone()
    if not order:
        flash('出库单不存在', 'danger')
        return redirect(url_for('sales_order_list'))

    db.execute("DELETE FROM sales_item_transport WHERE sales_item_id IN (SELECT id FROM sales_order_items WHERE sales_order_id=?)", (id,))
    db.execute("DELETE FROM sales_order_items WHERE sales_order_id = ?", (id,))
    db.execute("DELETE FROM sales_orders WHERE id = ?", (id,))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '删除销售出库单', f'出库单ID: {id}')
    flash('出库单已删除', 'success')
    return redirect(url_for('sales_order_list'))


@app.route('/api/sales/order/item/add', methods=['POST'])
@login_required
def api_sales_order_item_add():
    """API添加销售明细"""
    db = get_db()
    data = request.get_json() or request.form
    sales_order_id = data.get('sales_order_id', type=int)
    item_name = data.get('item_name', '')
    specification = data.get('specification', '')
    unit = data.get('unit', '')
    quantity = data.get('quantity', type=float) or 0
    unit_price = data.get('unit_price', type=float) or 0
    amount = round(quantity * unit_price, 2)

    max_order = db.execute("SELECT COALESCE(MAX(sort_order), 0) FROM sales_order_items WHERE sales_order_id=?", (sales_order_id,)).fetchone()[0]

    db.execute("""INSERT INTO sales_order_items (sales_order_id, item_name, specification, unit, quantity, unit_price, amount, sort_order)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
               (sales_order_id, item_name, specification, unit, quantity, unit_price, amount, max_order + 1))

    total_amount = db.execute("SELECT COALESCE(SUM(amount), 0) FROM sales_order_items WHERE sales_order_id=?", (sales_order_id,)).fetchone()[0]
    total_quantity = db.execute("SELECT COALESCE(SUM(quantity), 0) FROM sales_order_items WHERE sales_order_id=?", (sales_order_id,)).fetchone()[0]
    db.execute("UPDATE sales_orders SET total_amount=?, total_quantity=? WHERE id=?", (total_amount, total_quantity, sales_order_id))
    db.commit()

    item_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({'success': True, 'item_id': item_id, 'amount': amount,
                    'total_amount': total_amount, 'total_quantity': total_quantity})


@app.route('/api/sales/order/item/<int:item_id>/edit', methods=['POST'])
@login_required
def api_sales_order_item_edit(item_id):
    """API编辑销售明细"""
    db = get_db()
    data = request.get_json() or request.form
    item = db.execute("SELECT * FROM sales_order_items WHERE id=?", (item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': '明细不存在'}), 404

    item_name = data.get('item_name', item['item_name'])
    specification = data.get('specification', item['specification'])
    unit = data.get('unit', item['unit'])
    quantity = data.get('quantity', item['quantity'])
    unit_price = data.get('unit_price', item['unit_price'])
    amount = round(float(quantity) * float(unit_price), 2)

    db.execute("""UPDATE sales_order_items SET item_name=?, specification=?, unit=?, quantity=?, unit_price=?, amount=?
                  WHERE id=?""", (item_name, specification, unit, quantity, unit_price, amount, item_id))

    total_amount = db.execute("SELECT COALESCE(SUM(amount), 0) FROM sales_order_items WHERE sales_order_id=?", (item['sales_order_id'],)).fetchone()[0]
    total_quantity = db.execute("SELECT COALESCE(SUM(quantity), 0) FROM sales_order_items WHERE sales_order_id=?", (item['sales_order_id'],)).fetchone()[0]
    db.execute("UPDATE sales_orders SET total_amount=?, total_quantity=? WHERE id=?", (total_amount, total_quantity, item['sales_order_id']))
    db.commit()
    return jsonify({'success': True, 'amount': amount, 'total_amount': total_amount, 'total_quantity': total_quantity})


@app.route('/api/sales/order/item/<int:item_id>/delete', methods=['POST'])
@login_required
def api_sales_order_item_delete(item_id):
    """API删除销售明细"""
    db = get_db()
    item = db.execute("SELECT * FROM sales_order_items WHERE id=?", (item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': '明细不存在'}), 404

    sales_order_id = item['sales_order_id']
    db.execute("DELETE FROM sales_item_transport WHERE sales_item_id=?", (item_id,))
    db.execute("DELETE FROM sales_order_items WHERE id=?", (item_id,))

    total_amount = db.execute("SELECT COALESCE(SUM(amount), 0) FROM sales_order_items WHERE sales_order_id=?", (sales_order_id,)).fetchone()[0]
    total_quantity = db.execute("SELECT COALESCE(SUM(quantity), 0) FROM sales_order_items WHERE sales_order_id=?", (sales_order_id,)).fetchone()[0]
    db.execute("UPDATE sales_orders SET total_amount=?, total_quantity=? WHERE id=?", (total_amount, total_quantity, sales_order_id))
    db.commit()
    return jsonify({'success': True, 'total_amount': total_amount, 'total_quantity': total_quantity})


@app.route('/sales/order/<int:order_id>/item/import', methods=['GET', 'POST'])
@login_required
def sales_order_item_import(order_id):
    """销售明细导入"""
    db = get_db()
    order = db.execute("SELECT * FROM sales_orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        flash('出库单不存在', 'danger')
        return redirect(url_for('sales_order_list'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择文件', 'danger')
            return redirect(request.url)

        f = request.files['file']
        if not f.filename:
            flash('请选择文件', 'danger')
            return redirect(request.url)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            count = 0
            max_order = db.execute("SELECT COALESCE(MAX(sort_order), 0) FROM sales_order_items WHERE sales_order_id=?", (order_id,)).fetchone()[0]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                item_name = str(row[0]).strip() if row[0] else ''
                specification = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                unit = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                quantity = float(row[3]) if len(row) > 3 and row[3] else 0
                unit_price = float(row[4]) if len(row) > 4 and row[4] else 0
                amount = round(quantity * unit_price, 2)
                max_order += 1

                db.execute("""INSERT INTO sales_order_items (sales_order_id, item_name, specification, unit, quantity, unit_price, amount, sort_order)
                              VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                           (order_id, item_name, specification, unit, quantity, unit_price, amount, max_order))
                count += 1

            total_amount = db.execute("SELECT COALESCE(SUM(amount), 0) FROM sales_order_items WHERE sales_order_id=?", (order_id,)).fetchone()[0]
            total_quantity = db.execute("SELECT COALESCE(SUM(quantity), 0) FROM sales_order_items WHERE sales_order_id=?", (order_id,)).fetchone()[0]
            db.execute("UPDATE sales_orders SET total_amount=?, total_quantity=? WHERE id=?", (total_amount, total_quantity, order_id))
            db.commit()
            add_log(session.get('user_id'), session.get('username', ''), '导入销售明细', f'出库单ID: {order_id}, 导入{count}条')
            flash(f'成功导入 {count} 条明细', 'success')
        except Exception as e:
            flash(f'导入失败: {str(e)}', 'danger')

        return redirect(url_for('sales_order_detail', id=order_id))

    return render_template('sales_order_item_import.html', order=order)


@app.route('/sales/order/<int:order_id>/reconciliation')
@login_required
def sales_order_reconciliation(order_id):
    """对账差异分析"""
    db = get_db()
    order = db.execute("""SELECT so.*, p.name as project_name
                          FROM sales_orders so
                          LEFT JOIN projects p ON so.project_id = p.id
                          WHERE so.id = ?""", (order_id,)).fetchone()
    if not order:
        flash('出库单不存在', 'danger')
        return redirect(url_for('sales_order_list'))

    items = db.execute("""SELECT soi.*,
                                 COALESCE(sit_total.linked_qty, 0) as linked_qty,
                                 soi.quantity - COALESCE(sit_total.linked_qty, 0) as diff_qty
                          FROM sales_order_items soi
                          LEFT JOIN (SELECT sales_item_id, SUM(quantity) as linked_qty
                                     FROM sales_item_transport GROUP BY sales_item_id) sit_total
                          ON soi.id = sit_total.sales_item_id
                          WHERE soi.sales_order_id = ?
                          ORDER BY soi.sort_order, soi.id""", (order_id,)).fetchall()

    # 关联回款
    payments = db.execute("""SELECT * FROM sales_payments
                             WHERE customer_name = ? AND project_id = ?
                             ORDER BY payment_date DESC""",
                          (order['customer_name'], order['project_id'])).fetchall()

    total_payment = sum(p['amount'] for p in payments) if payments else 0
    diff_amount = order['total_amount'] - total_payment

    return render_template('sales_order_reconciliation.html', order=order, items=items,
                           payments=payments, total_payment=total_payment, diff_amount=diff_amount)


@app.route('/api/sales/batch/link/transport', methods=['POST'])
@login_required
def api_sales_batch_link_transport():
    """批量关联运输记录"""
    db = get_db()
    data = request.get_json() or request.form
    sales_item_id = data.get('sales_item_id', type=int)
    transport_ids = data.get('transport_ids', '')
    quantities = data.get('quantities', '')

    if not sales_item_id or not transport_ids:
        return jsonify({'success': False, 'message': '参数不完整'})

    item = db.execute("SELECT * FROM sales_order_items WHERE id=?", (sales_item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': '明细不存在'})

    tid_list = [int(x.strip()) for x in transport_ids.split(',') if x.strip()]
    qty_list = [float(x.strip()) for x in quantities.split(',') if x.strip()]

    count = 0
    for tid, qty in zip(tid_list, qty_list):
        if qty <= 0:
            continue
        # 检查是否已关联
        existing = db.execute("SELECT id FROM sales_item_transport WHERE sales_item_id=? AND transport_id=?",
                              (sales_item_id, tid)).fetchone()
        if existing:
            db.execute("UPDATE sales_item_transport SET quantity=? WHERE id=?", (qty, existing['id']))
        else:
            db.execute("INSERT INTO sales_item_transport (sales_item_id, transport_id, quantity) VALUES (?, ?, ?)",
                       (sales_item_id, tid, qty))
        count += 1

    db.commit()
    return jsonify({'success': True, 'count': count})


# ==================== 运输记录 ====================

def _parse_excel_cell_date(value):
    if value is None or value == '':
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()


@app.route('/transport/import', methods=['GET', 'POST'])
@login_required
def transport_import():
    """运费明细导入（Excel），带逐行容错与友好提示。"""
    db = get_db()
    if request.method == 'POST':
        import_type = (request.form.get('import_type') or 'excel').strip()
        purchase_id = request.form.get('purchase_id', type=int)
        back_url = (
            url_for('purchase_detail', id=purchase_id)
            if purchase_id else url_for('transport_import')
        )

        # 校验采购单
        if purchase_id:
            po = db.execute("SELECT id FROM purchase_orders WHERE id=?", (purchase_id,)).fetchone()
            if not po:
                flash('关联的采购单不存在，请重新选择', 'danger')
                return redirect(back_url)

        # 图片识别暂未在本入口实现，给出明确提示而非静默失败
        if import_type == 'ocr' or ('ocr_image' in request.files and 'file' not in request.files and 'excel_file' not in request.files):
            flash('运费图片识别功能尚未开放，请使用「Excel 导入」或「手动添加」录入运费明细', 'warning')
            return redirect(back_url)

        # 取文件（兼容两种字段名）
        f = request.files.get('excel_file') or request.files.get('file')
        if not f or not f.filename:
            flash('请选择要导入的 Excel 文件', 'danger')
            return redirect(back_url)

        if not f.filename.lower().endswith(('.xlsx', '.xls')):
            flash('文件格式不支持，请上传 .xlsx 或 .xls 文件', 'danger')
            return redirect(back_url)

        try:
            import openpyxl
        except ImportError:
            flash('服务器缺少 openpyxl 组件，无法解析 Excel，请联系管理员', 'danger')
            return redirect(back_url)

        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            flash(f'无法读取该 Excel 文件（可能已损坏或非标准格式）：{e}', 'danger')
            return redirect(back_url)

        ws = wb.active
        if ws is None or ws.max_row < 2:
            flash('Excel 中没有可导入的数据行（第 1 行为表头，数据请从第 2 行开始）', 'warning')
            return redirect(back_url)

        tr_cols = _table_columns(db, 'transport_records')
        tpi_cols = _table_columns(db, 'transport_purchase_items')
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        success = 0
        errors = []
        empty_skipped = 0
        # 列：车牌号、司机、运输日期、数量、单价运费、运费金额、备注
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(c is None or str(c).strip() == '' for c in row):
                empty_skipped += 1
                continue
            try:
                def cell(i):
                    return row[i] if len(row) > i else None

                vehicle_no = str(cell(0)).strip() if cell(0) is not None else ''
                driver_name = str(cell(1)).strip() if cell(1) is not None else ''
                transport_date = _parse_excel_cell_date(cell(2))
                quantity = _coerce_float(cell(3))
                unit_price = _coerce_float(cell(4))
                freight_amount = _coerce_float(cell(5), round(quantity * unit_price, 2))
                if freight_amount == 0 and quantity and unit_price:
                    freight_amount = round(quantity * unit_price, 2)
                remark = str(cell(6)).strip() if cell(6) is not None else ''

                if not vehicle_no and not driver_name and freight_amount == 0:
                    errors.append(f'第 {idx} 行：车牌号、司机、运费金额均为空，已跳过')
                    continue

                fields = {
                    'batch_no': f'第{success + 1}车',
                    'vehicle_no': vehicle_no,
                    'driver_name': driver_name,
                    'transport_date': transport_date or None,
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'freight_amount': freight_amount,
                    'remark': remark,
                }
                if 'purchase_id' in tr_cols:
                    fields['purchase_id'] = purchase_id
                if 'created_by' in tr_cols:
                    fields['created_by'] = session.get('user_id')
                if 'created_at' in tr_cols:
                    fields['created_at'] = now
                fields = {k: v for k, v in fields.items() if k in tr_cols}

                columns = ', '.join(fields.keys())
                placeholders = ', '.join('?' for _ in fields)
                cur = db.execute(
                    f"INSERT INTO transport_records ({columns}) VALUES ({placeholders})",
                    list(fields.values()),
                )
                success += 1
            except Exception as e:
                errors.append(f'第 {idx} 行：解析失败（{e}）')

        try:
            db.commit()
        except Exception as e:
            db.rollback()
            flash(f'导入写入失败，已回滚：{e}', 'danger')
            return redirect(back_url)

        add_log(session.get('user_id'), session.get('username', ''), '导入运输记录',
                f'采购单 {purchase_id or "-"}，成功 {success} 条')

        if success:
            msg = f'成功导入 {success} 条运费明细'
            if empty_skipped:
                msg += f'，跳过 {empty_skipped} 个空行'
            flash(msg, 'success')
        if errors:
            preview = '；'.join(errors[:5])
            more = f'（另有 {len(errors) - 5} 条问题未显示）' if len(errors) > 5 else ''
            flash(f'有 {len(errors)} 行未导入：{preview}{more}', 'warning')
        if not success and not errors:
            flash('未发现可导入的数据行，请检查 Excel 是否按模板填写', 'warning')

        return redirect(back_url)

    purchases = db.execute(
        "SELECT id, purchase_no, supplier FROM purchase_orders ORDER BY id DESC"
    ).fetchall()
    return render_template('transport_import.html', purchases=purchases)


@app.route('/api/transport/list')
@login_required
def api_transport_list():
    """运输记录列表API"""
    db = get_db()
    project_id = request.args.get('project_id', type=int)
    purchase_id = request.args.get('purchase_id', type=int)
    keyword = request.args.get('keyword', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    sql = """SELECT tr.*, p.name as project_name, po.purchase_no
             FROM transport_records tr
             LEFT JOIN projects p ON tr.project_id = p.id
             LEFT JOIN purchase_orders po ON tr.purchase_id = po.id
             WHERE 1=1"""
    params = []

    if project_id:
        sql += " AND tr.project_id = ?"
        params.append(project_id)
    if purchase_id:
        sql += " AND tr.purchase_id = ?"
        params.append(purchase_id)
    if keyword:
        sql += " AND (tr.batch_no LIKE ? OR tr.vehicle_no LIKE ? OR tr.driver_name LIKE ?)"
        params.extend([f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'])

    # 总数
    count_sql = sql.replace("SELECT tr.*, p.name as project_name, po.purchase_no", "SELECT COUNT(*)")
    total = db.execute(count_sql, params).fetchone()[0]

    sql += " ORDER BY tr.transport_date DESC, tr.id DESC"
    sql += f" LIMIT {per_page} OFFSET {(page - 1) * per_page}"

    records = db.execute(sql, params).fetchall()
    result = [dict(r) for r in records]

    return jsonify({'success': True, 'data': result, 'total': total,
                    'page': page, 'per_page': per_page,
                    'total_pages': (total + per_page - 1) // per_page})


@app.route('/api/transport/<int:id>')
@login_required
def api_transport_get(id):
    """运输记录详情API"""
    db = get_db()
    record = db.execute("""SELECT tr.*, p.name as project_name, po.purchase_no
                          FROM transport_records tr
                          LEFT JOIN projects p ON tr.project_id = p.id
                          LEFT JOIN purchase_orders po ON tr.purchase_id = po.id
                          WHERE tr.id = ?""", (id,)).fetchone()
    if not record:
        return jsonify({'success': False, 'message': '记录不存在'}), 404

    result = dict(record)

    # 关联的采购明细
    linked_items = db.execute("""SELECT tpi.*, pi.item_name, pi.specification, pi.unit
                                FROM transport_purchase_items tpi
                                JOIN purchase_items pi ON tpi.purchase_item_id = pi.id
                                WHERE tpi.transport_id = ?""", (id,)).fetchall()
    result['linked_purchase_items'] = [dict(r) for r in linked_items]

    # 关联的销售明细
    linked_sales = db.execute("""SELECT sit.*, soi.item_name, soi.specification, soi.unit
                                FROM sales_item_transport sit
                                JOIN sales_order_items soi ON sit.sales_item_id = soi.id
                                WHERE sit.transport_id = ?""", (id,)).fetchall()
    result['linked_sales_items'] = [dict(r) for r in linked_sales]

    return jsonify({'success': True, 'data': result})


@app.route('/api/transport/save', methods=['POST'])
@login_required
def api_transport_save():
    """保存运输记录（按实际表结构动态写入，兼容采购明细关联）"""
    db = get_db()
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        payload = request.form.to_dict()

    def _val(key, default=''):
        return (payload.get(key) if payload.get(key) is not None else default)

    record_id = payload.get('id')
    try:
        record_id = int(record_id) if record_id not in (None, '', 'null') else None
    except (TypeError, ValueError):
        record_id = None

    purchase_id = payload.get('purchase_id')
    try:
        purchase_id = int(purchase_id) if purchase_id not in (None, '', 'null') else None
    except (TypeError, ValueError):
        purchase_id = None

    quantity = _coerce_float(payload.get('quantity'))
    unit_price = _coerce_float(payload.get('unit_price'))
    freight_amount = _coerce_float(payload.get('freight_amount'), round(quantity * unit_price, 2))

    # 采购明细关联（多选）
    item_ids = []
    if isinstance(payload, dict):
        raw_ids = payload.get('purchase_item_ids[]') or payload.get('purchase_item_ids')
        if isinstance(raw_ids, str):
            item_ids = [raw_ids]
        elif isinstance(raw_ids, list):
            item_ids = raw_ids
    if not item_ids:
        item_ids = request.form.getlist('purchase_item_ids[]')
    item_ids = [int(i) for i in item_ids if str(i).strip().isdigit()]

    tr_cols = _table_columns(db, 'transport_records')

    fields = {
        'purchase_id': purchase_id,
        'batch_no': _val('batch_no'),
        'vehicle_no': _val('vehicle_no'),
        'driver_name': _val('driver_name'),
        'driver_phone': _val('driver_phone'),
        'transport_date': _val('transport_date') or None,
        'quantity': quantity,
        'unit_price': unit_price,
        'freight_amount': freight_amount,
        'remark': _val('remark'),
    }
    invoice_id = payload.get('invoice_id')
    try:
        invoice_id = int(invoice_id) if invoice_id not in (None, '', 'null') else None
    except (TypeError, ValueError):
        invoice_id = None
    if 'invoice_id' in tr_cols:
        fields['invoice_id'] = invoice_id
    if 'project_id' in tr_cols:
        proj = payload.get('project_id')
        try:
            fields['project_id'] = int(proj) if proj not in (None, '', 'null') else None
        except (TypeError, ValueError):
            fields['project_id'] = None
    if 'purchase_item_id' in tr_cols:
        fields['purchase_item_id'] = item_ids[0] if item_ids else None

    # 只保留实际存在的列
    fields = {k: v for k, v in fields.items() if k in tr_cols}

    if record_id:
        assignments = ', '.join(f'{k}=?' for k in fields)
        db.execute(
            f"UPDATE transport_records SET {assignments} WHERE id=?",
            list(fields.values()) + [record_id],
        )
    else:
        if 'created_by' in tr_cols:
            fields['created_by'] = session.get('user_id')
        if 'created_at' in tr_cols:
            fields['created_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        columns = ', '.join(fields.keys())
        placeholders = ', '.join('?' for _ in fields)
        cur = db.execute(
            f"INSERT INTO transport_records ({columns}) VALUES ({placeholders})",
            list(fields.values()),
        )
        record_id = cur.lastrowid

    # 维护采购明细关联表
    db.execute("DELETE FROM transport_purchase_items WHERE transport_id=?", (record_id,))
    tpi_cols = _table_columns(db, 'transport_purchase_items')
    for iid in item_ids:
        if 'quantity' in tpi_cols:
            db.execute(
                "INSERT INTO transport_purchase_items (transport_id, purchase_item_id, quantity) VALUES (?, ?, ?)",
                (record_id, iid, quantity if len(item_ids) == 1 else 0),
            )
        else:
            db.execute(
                "INSERT INTO transport_purchase_items (transport_id, purchase_item_id) VALUES (?, ?)",
                (record_id, iid),
            )

    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '保存运输记录', f'记录ID: {record_id}')
    return jsonify({'success': True, 'id': record_id})


# ==================== 合同管理 ====================

@app.route('/contract/list')
@login_required
def contract_list():
    """合同列表"""
    db = get_db()
    contract_type = request.args.get('contract_type', '')
    project_id = request.args.get('project_id', type=int)
    keyword = request.args.get('keyword', '')

    sql = """SELECT c.*, p.name as project_name
             FROM contracts c
             LEFT JOIN projects p ON c.project_id = p.id
             WHERE 1=1"""
    params = []

    if contract_type:
        sql += " AND c.contract_type = ?"
        params.append(contract_type)
    if project_id:
        sql += " AND c.project_id = ?"
        params.append(project_id)
    if keyword:
        sql += " AND (c.contract_no LIKE ? OR c.contract_name LIKE ? OR c.party LIKE ?)"
        params.extend([f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'])

    sql += " ORDER BY c.created_at DESC"
    contracts = db.execute(sql, params).fetchall()

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    return render_template('contract_list.html', contracts=contracts, projects=projects,
                           contract_type=contract_type, project_id=project_id, keyword=keyword)


@app.route('/contract/add', methods=['GET', 'POST'])
@login_required
def contract_add():
    """新增合同"""
    db = get_db()
    if request.method == 'POST':
        project_id = request.form.get('project_id', type=int)
        contract_no = request.form.get('contract_no', '')
        contract_name = request.form.get('contract_name', '')
        contract_type = request.form.get('contract_type', '')
        amount = request.form.get('amount', type=float) or 0
        tax_rate = request.form.get('tax_rate', type=float) or 0
        party = request.form.get('party', '')
        sign_date = request.form.get('sign_date', '') or None
        start_date = request.form.get('start_date', '') or None
        end_date = request.form.get('end_date', '') or None
        status = request.form.get('status', '执行中')
        remark = request.form.get('remark', '')

        # 处理附件
        attachment = ''
        if 'attachment' in request.files:
            f = request.files['attachment']
            if f and f.filename:
                ext = os.path.splitext(f.filename)[1]
                filename = f'contract_{uuid.uuid4().hex[:8]}{ext}'
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                attachment = filename

        db.execute("""INSERT INTO contracts (project_id, contract_no, contract_name, contract_type,
                      amount, tax_rate, party, sign_date, start_date, end_date, status, attachment, remark, created_by, created_at)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                   (project_id, contract_no, contract_name, contract_type, amount, tax_rate, party,
                    sign_date, start_date, end_date, status, attachment, remark,
                    session.get('user_id'), datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '新增合同', f'合同号: {contract_no}')
        flash('合同创建成功', 'success')
        return redirect(url_for('contract_list'))

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    return render_template('contract_form.html', projects=projects, now=datetime.now())


@app.route('/contract/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def contract_edit(id):
    """编辑合同"""
    db = get_db()
    contract = db.execute("SELECT * FROM contracts WHERE id = ?", (id,)).fetchone()
    if not contract:
        flash('合同不存在', 'danger')
        return redirect(url_for('contract_list'))

    if request.method == 'POST':
        project_id = request.form.get('project_id', type=int)
        contract_no = request.form.get('contract_no', '')
        contract_name = request.form.get('contract_name', '')
        contract_type = request.form.get('contract_type', '')
        amount = request.form.get('amount', type=float) or 0
        tax_rate = request.form.get('tax_rate', type=float) or 0
        party = request.form.get('party', '')
        sign_date = request.form.get('sign_date', '') or None
        start_date = request.form.get('start_date', '') or None
        end_date = request.form.get('end_date', '') or None
        status = request.form.get('status', '执行中')
        remark = request.form.get('remark', '')

        attachment = contract['attachment'] or ''
        if 'attachment' in request.files:
            f = request.files['attachment']
            if f and f.filename:
                ext = os.path.splitext(f.filename)[1]
                filename = f'contract_{uuid.uuid4().hex[:8]}{ext}'
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                attachment = filename

        db.execute("""UPDATE contracts SET project_id=?, contract_no=?, contract_name=?, contract_type=?,
                      amount=?, tax_rate=?, party=?, sign_date=?, start_date=?, end_date=?, status=?,
                      attachment=?, remark=?
                      WHERE id=?""",
                   (project_id, contract_no, contract_name, contract_type, amount, tax_rate, party,
                    sign_date, start_date, end_date, status, attachment, remark, id))
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '编辑合同', f'合同ID: {id}')
        flash('合同更新成功', 'success')
        return redirect(url_for('contract_list'))

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    return render_template('contract_form.html', contract=contract, projects=projects, edit_mode=True)


@app.route('/contract/<int:id>/delete', methods=['POST'])
@login_required
def contract_delete(id):
    """删除合同"""
    db = get_db()
    contract = db.execute("SELECT * FROM contracts WHERE id = ?", (id,)).fetchone()
    if not contract:
        flash('合同不存在', 'danger')
        return redirect(url_for('contract_list'))

    db.execute("DELETE FROM contracts WHERE id = ?", (id,))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '删除合同', f'合同ID: {id}')
    flash('合同已删除', 'success')
    return redirect(url_for('contract_list'))


# ==================== 对账管理 ====================

def _reconciliation_list_sql(db, project_id=None, status_filter=''):
    """对账列表查询（兼容 purchase_id 关联采购单的服务器表结构）。"""
    recon_cols = {r[1] for r in db.execute('PRAGMA table_info(reconciliations)').fetchall()}
    if 'purchase_id' in recon_cols:
        sql = """SELECT r.*, po.purchase_no, po.supplier, po.project_id,
                        p.name as project_name
                 FROM reconciliations r
                 LEFT JOIN purchase_orders po ON r.purchase_id = po.id
                 LEFT JOIN projects p ON po.project_id = p.id
                 WHERE 1=1"""
    else:
        sql = """SELECT r.*, p.name as project_name
                 FROM reconciliations r
                 LEFT JOIN projects p ON r.project_id = p.id
                 WHERE 1=1"""
    params = []
    if project_id:
        if 'purchase_id' in recon_cols:
            sql += ' AND po.project_id = ?'
        else:
            sql += ' AND r.project_id = ?'
        params.append(project_id)
    if status_filter:
        sql += ' AND r.status = ?'
        params.append(status_filter)
    sql += ' ORDER BY r.created_at DESC'
    return sql, params


def _pending_purchases_for_recon(db, project_id=None):
    """可对账但尚未生成对账单的采购单。"""
    po_cols = {r[1] for r in db.execute('PRAGMA table_info(purchase_orders)').fetchall()}
    date_col = 'purchase_date' if 'purchase_date' in po_cols else 'order_date'
    tr_cols = {r[1] for r in db.execute('PRAGMA table_info(transport_records)').fetchall()}
    transport_exists = (
        'EXISTS (SELECT 1 FROM transport_records tr WHERE tr.purchase_id = po.id)'
        if 'purchase_id' in tr_cols else
        """EXISTS (
            SELECT 1 FROM transport_purchase_items tpi
            JOIN purchase_items pi ON tpi.purchase_item_id = pi.id
            WHERE pi.purchase_id = po.id
        )"""
    )
    sql = f"""
        SELECT po.*, p.name as project_name
        FROM purchase_orders po
        LEFT JOIN projects p ON po.project_id = p.id
        WHERE po.status IN ('partial', 'delivered', 'completed', 'submitted', '已提交', '运输中')
          AND po.id NOT IN (
              SELECT purchase_id FROM reconciliations
              WHERE purchase_id IS NOT NULL
          )
          AND {transport_exists}
    """
    params = []
    if project_id:
        sql += ' AND po.project_id = ?'
        params.append(project_id)
    sql += f' ORDER BY po.{date_col} DESC, po.id DESC'
    return db.execute(sql, params).fetchall()


@app.route('/reconciliation/list')
@login_required
def reconciliation_list():
    """对账单列表"""
    db = get_db()
    project_id = request.args.get('project_id', type=int)
    status_filter = request.args.get('status', '')

    sql, params = _reconciliation_list_sql(db, project_id, status_filter)
    reconciliations = db.execute(sql, params).fetchall()
    pending_purchases = _pending_purchases_for_recon(db, project_id)
    projects = db.execute('SELECT id, name FROM projects ORDER BY name').fetchall()
    filters = {'project_id': project_id, 'status': status_filter}
    return render_template(
        'reconciliation_list.html',
        reconciliations=reconciliations,
        pending_purchases=pending_purchases,
        projects=projects,
        filters=filters,
        project_id=project_id,
        type=status_filter,
    )


@app.route('/reconciliation/<int:id>')
@login_required
def reconciliation_detail(id):
    """对账单详情"""
    db = get_db()
    recon_cols = {r[1] for r in db.execute('PRAGMA table_info(reconciliations)').fetchall()}
    if 'purchase_id' in recon_cols:
        recon = db.execute("""
            SELECT r.*, po.purchase_no, po.supplier, po.project_id,
                   p.name as project_name
            FROM reconciliations r
            LEFT JOIN purchase_orders po ON r.purchase_id = po.id
            LEFT JOIN projects p ON po.project_id = p.id
            WHERE r.id = ?
        """, (id,)).fetchone()
    else:
        recon = db.execute("""
            SELECT r.*, p.name as project_name
            FROM reconciliations r
            LEFT JOIN projects p ON r.project_id = p.id
            WHERE r.id = ?
        """, (id,)).fetchone()
    if not recon:
        flash('对账单不存在', 'danger')
        return redirect(url_for('reconciliation_list'))

    items = []
    transports = []
    purchase_id = recon['purchase_id'] if 'purchase_id' in recon_cols else None
    if purchase_id:
        items = db.execute(
            'SELECT * FROM purchase_items WHERE purchase_id=? ORDER BY sort_order, id',
            (purchase_id,),
        ).fetchall()
        tr_cols = {r[1] for r in db.execute('PRAGMA table_info(transport_records)').fetchall()}
        inv_cols = {r[1] for r in db.execute('PRAGMA table_info(invoices)').fetchall()}
        inv_join = ' LEFT JOIN invoices i ON tr.invoice_id = i.id' if 'invoice_id' in tr_cols and inv_cols else ''
        inv_sel = ', i.invoice_no' if 'invoice_id' in tr_cols and inv_cols else ", NULL as invoice_no"
        item_join = ''
        item_sel = ", NULL as item_name"
        if 'purchase_item_id' in tr_cols:
            item_join = ' LEFT JOIN purchase_items pi ON tr.purchase_item_id = pi.id'
            item_sel = ', pi.item_name'
        if 'purchase_id' in tr_cols:
            transports = db.execute(f"""
                SELECT tr.*{inv_sel}{item_sel}
                FROM transport_records tr{inv_join}{item_join}
                WHERE tr.purchase_id = ?
                ORDER BY tr.transport_date, tr.id
            """, (purchase_id,)).fetchall()

    return render_template(
        'reconciliation_detail.html',
        recon=recon,
        items=items,
        transports=transports,
    )


# ==================== 发票管理 ====================

@app.route('/invoice/list')
@login_required
def invoice_list():
    """发票列表"""
    db = get_db()
    project_id = request.args.get('project_id', type=int)
    invoice_type = request.args.get('invoice_type', '')
    status = request.args.get('status', '')

    sql = """SELECT i.*, i.invoice_no as invoice_number,
                    p.name as project_name, c.contract_name
             FROM invoices i
             LEFT JOIN projects p ON i.project_id = p.id
             LEFT JOIN contracts c ON i.contract_id = c.id
             WHERE 1=1"""
    params = []

    if project_id:
        sql += " AND i.project_id = ?"
        params.append(project_id)
    if invoice_type:
        sql += " AND i.invoice_type = ?"
        params.append(invoice_type)
    if status:
        sql += " AND i.status = ?"
        params.append(status)

    sql += " ORDER BY i.created_at DESC"
    invoices = db.execute(sql, params).fetchall()
    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    return render_template('invoice_list.html', invoices=invoices, projects=projects,
                           project_id=project_id, invoice_type=invoice_type, status=status)


@app.route('/invoice/add', methods=['GET', 'POST'])
@login_required
def invoice_add():
    """新增发票"""
    db = get_db()
    if request.method == 'POST':
        project_id = request.form.get('project_id', type=int)
        contract_id = request.form.get('contract_id', type=int) or None
        invoice_no = request.form.get('invoice_no', '')
        invoice_type = request.form.get('invoice_type', '')
        amount = request.form.get('amount', type=float) or 0
        tax_rate = request.form.get('tax_rate', type=float) or 0
        tax_amount = round(amount * tax_rate / 100, 2) if tax_rate else 0
        remark = request.form.get('remark', '')

        db.execute("""INSERT INTO invoices (project_id, contract_id, invoice_no, invoice_type,
                      amount, tax_rate, tax_amount, remark)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                   (project_id, contract_id, invoice_no, invoice_type, amount, tax_rate, tax_amount, remark))
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''), '新增发票', f'发票号: {invoice_no}')
        flash('发票添加成功', 'success')
        return redirect(url_for('invoice_list'))

    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    contracts = db.execute("SELECT id, contract_name FROM contracts ORDER BY contract_name").fetchall()
    return render_template('invoice_form.html', projects=projects, contracts=contracts, now=datetime.now())


# ==================== 基础资料 ====================

@app.route('/base/data')
@login_required
def base_data():
    """基础资料首页"""
    db = get_db()
    supplier_count = db.execute("SELECT COUNT(*) FROM suppliers WHERE is_active=1").fetchone()[0]
    customer_count = db.execute("SELECT COUNT(*) FROM customers WHERE is_active=1").fetchone()[0]
    category_count = db.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
    payment_type_count = db.execute("SELECT COUNT(*) FROM payment_types").fetchone()[0]
    return render_template('base_data.html',
                           supplier_count=supplier_count,
                           customer_count=customer_count,
                           category_count=category_count,
                           payment_type_count=payment_type_count)


@app.route('/supplier/list')
@login_required
def supplier_list():
    """供应商列表"""
    db = get_db()
    keyword = request.args.get('keyword', '')
    suppliers = db.execute("SELECT * FROM suppliers WHERE is_active=1 AND name LIKE ? ORDER BY name",
                           (f'%{keyword}%',)).fetchall()
    return render_template('supplier_list.html', suppliers=suppliers, keyword=keyword)


@app.route('/supplier/add', methods=['GET', 'POST'])
@login_required
def supplier_add():
    """新增供应商"""
    if request.method == 'POST':
        return supplier_add_post()
    return render_template('supplier_form.html')


@app.route('/supplier/add', methods=['POST'])
@login_required
def supplier_add_post():
    """新增供应商 - POST处理"""
    db = get_db()
    name = request.form.get('name', '')
    contact = request.form.get('contact', '')
    phone = request.form.get('phone', '')
    email = request.form.get('email', '')
    address = request.form.get('address', '')
    delivery_address = request.form.get('delivery_address', '')
    bank_name = request.form.get('bank_name', '')
    bank_account = request.form.get('bank_account', '')
    bank_code = request.form.get('bank_code', '')
    tax_no = request.form.get('tax_no', '')
    invoice_title = request.form.get('invoice_title', '')
    invoice_addr_phone = request.form.get('invoice_addr_phone', '')
    invoice_bank_account = request.form.get('invoice_bank_account', '')
    tax_rate = request.form.get('tax_rate', type=float) or 0
    remark = request.form.get('remark', '')

    db.execute("""INSERT INTO suppliers (name, contact, phone, email, address, delivery_address,
                  bank_name, bank_account, bank_code, tax_no, invoice_title, invoice_addr_phone,
                  invoice_bank_account, tax_rate, remark, created_at, is_active)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
               (name, contact, phone, email, address, delivery_address, bank_name, bank_account,
                bank_code, tax_no, invoice_title, invoice_addr_phone, invoice_bank_account,
                tax_rate, remark, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '新增供应商', f'供应商: {name}')
    flash('供应商添加成功', 'success')
    return redirect(url_for('supplier_list'))


@app.route('/supplier/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def supplier_edit(id):
    """编辑供应商"""
    if request.method == 'POST':
        return supplier_edit_post(id)
    db = get_db()
    supplier = db.execute("SELECT * FROM suppliers WHERE id=?", (id,)).fetchone()
    if not supplier:
        flash('供应商不存在', 'danger')
        return redirect(url_for('supplier_list'))
    return render_template('supplier_form.html', supplier=supplier, edit_mode=True)


@app.route('/supplier/<int:id>/edit', methods=['POST'])
@login_required
def supplier_edit_post(id):
    """编辑供应商 - POST处理"""
    db = get_db()
    supplier = db.execute("SELECT * FROM suppliers WHERE id=?", (id,)).fetchone()
    if not supplier:
        flash('供应商不存在', 'danger')
        return redirect(url_for('supplier_list'))

    name = request.form.get('name', supplier['name'])
    contact = request.form.get('contact', supplier['contact'])
    phone = request.form.get('phone', supplier['phone'])
    email = request.form.get('email', supplier['email'])
    address = request.form.get('address', supplier['address'])
    delivery_address = request.form.get('delivery_address', supplier['delivery_address'])
    bank_name = request.form.get('bank_name', supplier['bank_name'])
    bank_account = request.form.get('bank_account', supplier['bank_account'])
    bank_code = request.form.get('bank_code', supplier['bank_code'])
    tax_no = request.form.get('tax_no', supplier['tax_no'])
    invoice_title = request.form.get('invoice_title', supplier['invoice_title'])
    invoice_addr_phone = request.form.get('invoice_addr_phone', supplier['invoice_addr_phone'])
    invoice_bank_account = request.form.get('invoice_bank_account', supplier['invoice_bank_account'])
    tax_rate = request.form.get('tax_rate', type=float) or 0
    remark = request.form.get('remark', supplier['remark'])

    db.execute("""UPDATE suppliers SET name=?, contact=?, phone=?, email=?, address=?, delivery_address=?,
                  bank_name=?, bank_account=?, bank_code=?, tax_no=?, invoice_title=?, invoice_addr_phone=?,
                  invoice_bank_account=?, tax_rate=?, remark=?
                  WHERE id=?""",
               (name, contact, phone, email, address, delivery_address, bank_name, bank_account,
                bank_code, tax_no, invoice_title, invoice_addr_phone, invoice_bank_account,
                tax_rate, remark, id))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '编辑供应商', f'供应商ID: {id}')
    flash('供应商更新成功', 'success')
    return redirect(url_for('supplier_list'))


def _partner_import_config(entity):
    """entity: supplier | customer"""
    if entity == 'customer':
        return {
            'table': 'customers',
            'role': 'customer',
            'entity_label': '客户',
            'list_url': 'customer_list',
            'ocr_url': 'customer_ocr_import',
            'excel_url': 'customer_excel_import',
            'upload_subdir': 'customers',
            'log_action': '客户',
        }
    return {
        'table': 'suppliers',
        'role': 'supplier',
        'entity_label': '供应商',
        'list_url': 'supplier_list',
        'ocr_url': 'supplier_ocr_import',
        'excel_url': 'supplier_excel_import',
        'upload_subdir': 'suppliers',
        'log_action': '供应商',
    }


def _handle_partner_ocr_import(entity):
    from partner_import_utils import recognize_company_info, save_partner_record
    cfg = _partner_import_config(entity)
    upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], cfg['upload_subdir'], 'ocr')
    os.makedirs(upload_dir, exist_ok=True)
    update_existing = request.form.get('update_existing') == '1'
    files = request.files.getlist('images')
    results = []
    stats = {'created': 0, 'updated': 0, 'skipped': 0, 'failed': 0}
    db = get_db()

    for f in files:
        if not f or not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower() or '.jpg'
        save_name = f"{cfg['upload_subdir']}_{uuid.uuid4().hex[:10]}{ext}"
        path = os.path.join(upload_dir, save_name)
        f.save(path)
        try:
            info = recognize_company_info(path, role=cfg['role'])
            if info.get('error'):
                stats['failed'] += 1
                results.append({
                    'filename': f.filename, 'status': 'failed',
                    'error': info['error'], 'message': '识别失败'
                })
                continue
            status = save_partner_record(db, cfg['table'], info, update_existing)
            stats[status] = stats.get(status, 0) + 1
            msg_map = {'created': '已新增', 'updated': '已更新', 'skipped': '已存在跳过'}
            results.append({
                'filename': f.filename, 'status': status,
                'data': info, 'message': msg_map.get(status, status)
            })
        except Exception as ex:
            stats['failed'] += 1
            results.append({
                'filename': f.filename, 'status': 'failed',
                'error': str(ex), 'message': '处理异常'
            })

    if request.method == 'POST' and files:
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''),
                f'{cfg["log_action"]}图片导入',
                f'新增{stats["created"]} 更新{stats["updated"]} 失败{stats["failed"]}',
                request.remote_addr)
        if stats['created'] or stats['updated']:
            flash(f'导入完成：新增 {stats["created"]}，更新 {stats["updated"]}', 'success')
        elif stats['failed']:
            flash('识别完成，但未能成功导入任何记录', 'warning')
        else:
            flash('所有记录均已存在，未新增', 'info')

    return render_template(
        'partner_ocr_import.html',
        entity_label=cfg['entity_label'],
        back_url=url_for(cfg['list_url']),
        results=results if request.method == 'POST' else None,
        stats=stats if request.method == 'POST' else None,
    )


def _handle_partner_excel_import(entity):
    from partner_import_utils import parse_partner_excel, save_partner_record
    cfg = _partner_import_config(entity)
    stats = None
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename:
            flash('请选择 Excel 文件', 'warning')
            return redirect(url_for(cfg['excel_url']))
        upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], cfg['upload_subdir'])
        os.makedirs(upload_dir, exist_ok=True)
        path = os.path.join(upload_dir, f'import_{uuid.uuid4().hex[:8]}.xlsx')
        f.save(path)
        records, err = parse_partner_excel(path)
        if err:
            flash(f'解析 Excel 失败: {err}', 'danger')
            return redirect(url_for(cfg['excel_url']))
        update_existing = request.form.get('update_existing') == '1'
        db = get_db()
        stats = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
        for i, row in enumerate(records, start=2):
            try:
                status = save_partner_record(db, cfg['table'], row, update_existing)
                stats[status] = stats.get(status, 0) + 1
            except Exception as ex:
                stats['errors'].append(f'第{i}行 {row.get("name","")}: {ex}')
        db.commit()
        add_log(session.get('user_id'), session.get('username', ''),
                f'{cfg["log_action"]}Excel导入',
                f'新增{stats["created"]} 更新{stats["updated"]}',
                request.remote_addr)
        flash(f'导入完成：新增 {stats["created"]}，更新 {stats["updated"]}，跳过 {stats["skipped"]}', 'success')

    return render_template(
        'partner_excel_import.html',
        entity_label=cfg['entity_label'],
        back_url=url_for(cfg['list_url']),
        stats=stats,
    )


@app.route('/supplier/import/ocr', methods=['GET', 'POST'])
@login_required
def supplier_ocr_import():
    return _handle_partner_ocr_import('supplier')


@app.route('/supplier/import/excel', methods=['GET', 'POST'])
@login_required
def supplier_excel_import():
    return _handle_partner_excel_import('supplier')


@app.route('/customer/import/ocr', methods=['GET', 'POST'])
@login_required
def customer_ocr_import():
    return _handle_partner_ocr_import('customer')


@app.route('/customer/import/excel', methods=['GET', 'POST'])
@login_required
def customer_excel_import():
    return _handle_partner_excel_import('customer')


@app.route('/supplier/<int:id>/delete', methods=['POST'])
@login_required
def supplier_delete(id):
    """删除供应商"""
    db = get_db()
    supplier = db.execute("SELECT * FROM suppliers WHERE id=?", (id,)).fetchone()
    if not supplier:
        flash('供应商不存在', 'danger')
        return redirect(url_for('supplier_list'))

    db.execute("UPDATE suppliers SET is_active=0 WHERE id=?", (id,))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '删除供应商', f'供应商ID: {id}')
    flash('供应商已删除', 'success')
    return redirect(url_for('supplier_list'))


@app.route('/customer/list')
@login_required
def customer_list():
    """客户列表"""
    db = get_db()
    keyword = request.args.get('keyword', '')
    customers = db.execute("SELECT * FROM customers WHERE is_active=1 AND name LIKE ? ORDER BY name",
                           (f'%{keyword}%',)).fetchall()
    return render_template('customer_list.html', customers=customers, keyword=keyword)


@app.route('/customer/add', methods=['GET', 'POST'])
@login_required
def customer_add():
    """新增客户"""
    if request.method == 'POST':
        return customer_add_post()
    return render_template('customer_form.html')


@app.route('/customer/add', methods=['POST'])
@login_required
def customer_add_post():
    """新增客户 - POST处理"""
    db = get_db()
    name = request.form.get('name', '')
    contact = request.form.get('contact', '')
    phone = request.form.get('phone', '')
    email = request.form.get('email', '')
    address = request.form.get('address', '')
    delivery_address = request.form.get('delivery_address', '')
    bank_name = request.form.get('bank_name', '')
    bank_account = request.form.get('bank_account', '')
    bank_code = request.form.get('bank_code', '')
    tax_no = request.form.get('tax_no', '')
    invoice_title = request.form.get('invoice_title', '')
    invoice_addr_phone = request.form.get('invoice_addr_phone', '')
    invoice_bank_account = request.form.get('invoice_bank_account', '')
    tax_rate = request.form.get('tax_rate', type=float) or 0
    remark = request.form.get('remark', '')

    db.execute("""INSERT INTO customers (name, contact, phone, email, address, delivery_address,
                  bank_name, bank_account, bank_code, tax_no, invoice_title, invoice_addr_phone,
                  invoice_bank_account, tax_rate, remark, created_at, is_active)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
               (name, contact, phone, email, address, delivery_address, bank_name, bank_account,
                bank_code, tax_no, invoice_title, invoice_addr_phone, invoice_bank_account,
                tax_rate, remark, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '新增客户', f'客户: {name}')
    flash('客户添加成功', 'success')
    return redirect(url_for('customer_list'))


@app.route('/customer/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def customer_edit(id):
    """编辑客户"""
    if request.method == 'POST':
        return customer_edit_post(id)
    db = get_db()
    customer = db.execute("SELECT * FROM customers WHERE id=?", (id,)).fetchone()
    if not customer:
        flash('客户不存在', 'danger')
        return redirect(url_for('customer_list'))
    return render_template('customer_form.html', customer=customer, edit_mode=True)


@app.route('/customer/<int:id>/edit', methods=['POST'])
@login_required
def customer_edit_post(id):
    """编辑客户 - POST处理"""
    db = get_db()
    customer = db.execute("SELECT * FROM customers WHERE id=?", (id,)).fetchone()
    if not customer:
        flash('客户不存在', 'danger')
        return redirect(url_for('customer_list'))

    name = request.form.get('name', customer['name'])
    contact = request.form.get('contact', customer['contact'])
    phone = request.form.get('phone', customer['phone'])
    email = request.form.get('email', customer['email'])
    address = request.form.get('address', customer['address'])
    delivery_address = request.form.get('delivery_address', customer['delivery_address'])
    bank_name = request.form.get('bank_name', customer['bank_name'])
    bank_account = request.form.get('bank_account', customer['bank_account'])
    bank_code = request.form.get('bank_code', customer['bank_code'])
    tax_no = request.form.get('tax_no', customer['tax_no'])
    invoice_title = request.form.get('invoice_title', customer['invoice_title'])
    invoice_addr_phone = request.form.get('invoice_addr_phone', customer['invoice_addr_phone'])
    invoice_bank_account = request.form.get('invoice_bank_account', customer['invoice_bank_account'])
    tax_rate = request.form.get('tax_rate', type=float) or 0
    remark = request.form.get('remark', customer['remark'])

    db.execute("""UPDATE customers SET name=?, contact=?, phone=?, email=?, address=?, delivery_address=?,
                  bank_name=?, bank_account=?, bank_code=?, tax_no=?, invoice_title=?, invoice_addr_phone=?,
                  invoice_bank_account=?, tax_rate=?, remark=?
                  WHERE id=?""",
               (name, contact, phone, email, address, delivery_address, bank_name, bank_account,
                bank_code, tax_no, invoice_title, invoice_addr_phone, invoice_bank_account,
                tax_rate, remark, id))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '编辑客户', f'客户ID: {id}')
    flash('客户更新成功', 'success')
    return redirect(url_for('customer_list'))


@app.route('/customer/<int:id>/delete', methods=['POST'])
@login_required
def customer_delete(id):
    """删除客户"""
    db = get_db()
    customer = db.execute("SELECT * FROM customers WHERE id=?", (id,)).fetchone()
    if not customer:
        flash('客户不存在', 'danger')
        return redirect(url_for('customer_list'))

    db.execute("UPDATE customers SET is_active=0 WHERE id=?", (id,))
    db.commit()
    add_log(session.get('user_id'), session.get('username', ''), '删除客户', f'客户ID: {id}')
    flash('客户已删除', 'success')
    return redirect(url_for('customer_list'))


@app.route('/payment_type/list')
@login_required
def payment_type_list():
    """付款类型管理"""
    db = get_db()
    types = db.execute("SELECT * FROM payment_types ORDER BY id").fetchall()
    return render_template('payment_type_list.html', payment_types=types)


# ==================== 文件服务 ====================

def resolve_attachment_path(filename):
    """解析附件在 UPLOAD_FOLDER 下的相对路径（兼容 OCR 子目录）。"""
    if not filename:
        return None
    rel = filename.replace('\\', '/').strip().lstrip('/')
    if rel.startswith('uploads/'):
        rel = rel[len('uploads/'):]
    base = app.config['UPLOAD_FOLDER']
    candidates = [rel]
    if '/' not in rel:
        candidates.append(f'ocr/{rel}')
    if rel.startswith('ocr/'):
        candidates.append(rel[4:])
    for path in candidates:
        if os.path.isfile(os.path.join(base, path)):
            return path
    return rel


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """提供上传文件访问"""
    rel = resolve_attachment_path(filename)
    return send_from_directory(app.config['UPLOAD_FOLDER'], rel)


# ==================== 批量关联 ====================

@app.route('/api/purchase/batch/link/transport', methods=['POST'])
@login_required
def api_purchase_batch_link_transport():
    """采购明细批量关联运输"""
    db = get_db()
    data = request.get_json() or request.form
    purchase_item_id = data.get('purchase_item_id', type=int)
    transport_ids = data.get('transport_ids', '')
    quantities = data.get('quantities', '')

    if not purchase_item_id or not transport_ids:
        return jsonify({'success': False, 'message': '参数不完整'})

    item = db.execute("SELECT * FROM purchase_items WHERE id=?", (purchase_item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': '明细不存在'})

    tid_list = [int(x.strip()) for x in transport_ids.split(',') if x.strip()]
    qty_list = [float(x.strip()) for x in quantities.split(',') if x.strip()]

    count = 0
    for tid, qty in zip(tid_list, qty_list):
        if qty <= 0:
            continue
        existing = db.execute("SELECT id FROM transport_purchase_items WHERE purchase_item_id=? AND transport_id=?",
                              (purchase_item_id, tid)).fetchone()
        if existing:
            db.execute("UPDATE transport_purchase_items SET quantity=? WHERE id=?", (qty, existing['id']))
        else:
            db.execute("INSERT INTO transport_purchase_items (transport_id, purchase_item_id, quantity) VALUES (?, ?, ?)",
                       (tid, purchase_item_id, qty))
        count += 1

    db.commit()
    return jsonify({'success': True, 'count': count})


# ==================== API：项目关联数据 ====================

@app.route('/api/project/<int:pid>/categories/grouped')
@login_required
def api_project_categories_grouped(pid):
    db = get_db()
    ensure_project_categories_table(db)
    return jsonify(get_grouped_categories_json(db, pid))


@app.route('/api/project/<int:pid>/config')
@login_required
def api_project_config(pid):
    db = get_db()
    p = db.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    freight = False
    if p and 'freight_link_enabled' in p.keys():
        freight = bool(p['freight_link_enabled'])
    return jsonify({'freight_link_enabled': freight})


@app.route('/api/project/<int:pid>/invoices')
@login_required
def api_project_invoices(pid):
    db = get_db()
    rows = db.execute(
        "SELECT id, invoice_no, invoice_type, amount FROM invoices "
        "WHERE project_id=? ORDER BY id DESC",
        (pid,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/project/<int:pid>/contracts')
@login_required
def api_project_contracts(pid):
    db = get_db()
    rows = db.execute(
        "SELECT id, contract_name, contract_type FROM contracts "
        "WHERE project_id=? ORDER BY contract_name",
        (pid,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/ocr/recognize', methods=['POST'])
@login_required
def api_ocr_recognize():
    """费用录入页：上传图片即时 OCR"""
    if 'image' not in request.files:
        return jsonify({'success': False, 'message': '请选择图片'}), 400
    file = request.files['image']
    if not file.filename:
        return jsonify({'success': False, 'message': '请选择图片'}), 400
    allowed = ('.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp')
    if not file.filename.lower().endswith(allowed):
        return jsonify({'success': False, 'message': '不支持的图片格式'}), 400

    filename = f"ocr_{uuid.uuid4().hex[:12]}_{file.filename}"
    ocr_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr')
    os.makedirs(ocr_dir, exist_ok=True)
    filepath = os.path.join(ocr_dir, filename)
    file.save(filepath)

    try:
        transactions, raw_text = ocr_recognize(filepath)
        if '无法识别文字' in raw_text or '请安装' in raw_text:
            return jsonify({
                'success': False,
                'message': raw_text,
                'raw_text': raw_text,
                'transactions': [],
            })
        return jsonify({
            'success': True,
            'transactions': transactions,
            'raw_text': raw_text,
            'count': len(transactions),
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/project/<int:pid>/purchases')
@login_required
def api_project_purchases(pid):
    db = get_db()
    rows = db.execute(
        "SELECT id, purchase_no, supplier FROM purchase_orders "
        "WHERE project_id=? ORDER BY id DESC",
        (pid,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


register_auth_hooks(app)

from reports_routes import register_reports_blueprint
register_reports_blueprint(app, get_db)

from route_extensions import register_missing_routes
register_missing_routes(app, {
    'login_required': login_required,
    'admin_required': admin_required,
    'get_db': get_db,
    'add_log': add_log,
    'datetime': datetime,
    'redirect': redirect,
    'url_for': url_for,
    'render_template': render_template,
    'request': request,
    'session': session,
    'flash': flash,
    'send_from_directory': send_from_directory,
    'os': os,
    'uuid': uuid,
    'recalc_investment_ratios': recalc_investment_ratios,
})


if __name__ == '__main__':
    init_db()
    # 启动定时备份线程
    scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
    scheduler_thread.start()
    print("=" * 50)
    print("  项目费用归集系统已启动")
    print("  默认账号: admin / admin123")
    print("  默认端口: 5002")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5002, debug=True)
